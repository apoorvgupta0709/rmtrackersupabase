"""Write the August schedules into a copy of the master workbook, for the owner.

A deliverable for a person to open in Excel, not an input for the pipeline — the
pipeline keeps reading the untouched master plus `schedule_supplement.xlsx`.

Two things the sheet does for itself, which this must not take over:

`SCHEDULE IN MT` is a formula over ACTUAL OD, TICKNESS, LENGTH and SCHEDULE in nos.
Only the quantity is written; the tonnage recalculates when the file opens. Writing a
value there would replace the workbook's own arithmetic with mine on some rows and not
others, which is worse than either alone.

NMPL orders the same size from two plants, so two August lines can land on one sheet
row. Those are summed. Overwriting would silently drop one plant's demand — 9 rows on
this month's build.
"""
import shutil
import warnings
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

SRC = Path("aug/raw/aug0826 rm tracker v1.xlsx")
OUT = Path("aug/aug0826 rm tracker v1 - August schedules.xlsx")
SHEET = "Schedule July"
HEADER_ROW, FIRST_DATA_ROW, LAST_DATA_ROW = 3, 4, 745

shutil.copy(SRC, OUT)
book = load_workbook(OUT)
ws = book[SHEET]

columns = {}
for cell in ws[HEADER_ROW]:
    name = str(cell.value).strip() if cell.value is not None else ""
    if name and name not in columns:
        columns[name] = cell.column

matched = pd.read_csv("aug/matched.csv")
supplement = pd.read_excel("aug/schedule_supplement.xlsx")

# 1. Existing rows, summing the lines that share one.
totals, notes = defaultdict(float), defaultdict(list)
for _, line in matched.iterrows():
    row = int(line["_row"])
    totals[row] += float(line["nos"])
    notes[row].append(str(line["source_note"]))
for row, pieces in totals.items():
    excel_row = FIRST_DATA_ROW + row
    assert excel_row <= LAST_DATA_ROW, excel_row
    ws.cell(excel_row, columns["SCHEDULE in nos"]).value = pieces
    if "Customer remarks" in columns:
        ws.cell(excel_row, columns["Customer remarks"]).value = (
            "Aug 2026 schedule · " + " + ".join(sorted(set(notes[row]))))

# 2. New sizes and the carried-forward customers, appended below the existing rows.
#    Their tonnage gets the same formula the sheet uses, written with plain cell
#    references so it does not depend on the row joining the sheet's Excel table.
def mt_formula(excel_row):
    od = get_column_letter(columns["ACTUAL OD"])
    thk = get_column_letter(columns["TICKNESS"])
    length = get_column_letter(columns["LENGTH"])
    nos = get_column_letter(columns["SCHEDULE in nos"])
    return (f"=(3.14*({od}{excel_row}-{thk}{excel_row})*{thk}{excel_row}"
            f"*{length}{excel_row}*7.85)/1000000*{nos}{excel_row}/1000")


appended = supplement[
    supplement["Customer remarks"].astype(str).str.contains("new size|Carried over", na=False)
]
cursor = LAST_DATA_ROW + 1
for _, line in appended.iterrows():
    for name, column in columns.items():
        if name in ("SCHEDULE IN MT", "DISPATCH"):
            continue
        if name in line.index and pd.notna(line[name]):
            value = line[name]
            ws.cell(cursor, column).value = (
                float(value) if isinstance(value, (int, float)) else str(value))
    ws.cell(cursor, columns["SCHEDULE IN MT"]).value = mt_formula(cursor)
    cursor += 1

book.save(OUT)
print(f"{len(totals)} existing rows given an August quantity "
      f"({len(matched)} lines, {len(matched) - len(totals)} of them summed onto a shared row)")
print(f"{len(appended)} rows appended, sheet now ends at Excel row {cursor - 1}")
print(f"sheets preserved: {len(book.sheetnames)} of {len(load_workbook(SRC).sheetnames)}")

# 3. Prove the quantities match what the dashboard published.
check = pd.read_excel(OUT, sheet_name=SHEET, header=2).iloc[:cursor - FIRST_DATA_ROW]
qty = pd.to_numeric(check["SCHEDULE in nos"], errors="coerce").fillna(0)
active = check[qty > 0]
print(f"\n{len(active)} active rows, {active['Helper Customer'].nunique()} customers")
mine = supplement.groupby("Helper Customer")["SCHEDULE in nos"].sum().round(0)
theirs = active.groupby("Helper Customer")["SCHEDULE in nos"].apply(
    lambda x: pd.to_numeric(x, errors="coerce").sum()).round(0)
print("\npieces per customer, supplement vs workbook:")
for name in sorted(mine.index):
    a, b = float(mine.get(name, 0)), float(theirs.get(name, 0))
    print(f"  {name[:32]:34}{a:>14,.0f}{b:>14,.0f}  {'ok' if abs(a - b) < 1 else 'MISMATCH'}")
