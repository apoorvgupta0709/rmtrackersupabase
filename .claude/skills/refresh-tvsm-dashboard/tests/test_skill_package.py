from __future__ import annotations

import hashlib
import importlib.util
import json
import re
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


SKILL_ROOT = Path(__file__).resolve().parents[1]
# The skill lives at `.claude/skills/<name>/`, so the repository root is three levels
# up, not one. Deriving it by counting parents broke silently when the package moved
# under `.claude/`: the checks that keep dumps/ out of the public deployment went on
# passing while looking inside `.claude/` for a `.vercelignore` that is not there.
REPO_ROOT = SKILL_ROOT.parents[2]


def vercelignore_patterns():
    """The patterns in `.vercelignore`, with the leading slash normalised away.

    Anchoring is not cosmetic and these tests must not push back against it. Unanchored,
    gitignore semantics match a directory of that name at *any* depth, so `dumps/` also
    excluded `lib/dumps/` and `supabase/` also excluded `lib/supabase/` — between them
    the whole of `lib/`. Every deployment then died with `Module not found` while the
    local build stayed green, because the files were missing only from the upload.

    What these tests care about is that the directory is excluded, not how it is spelt,
    so compare on the normalised form and let either spelling pass.
    """
    text = (REPO_ROOT / ".vercelignore").read_text(encoding="utf-8")
    return {
        line.strip().lstrip("/")
        for line in text.splitlines()
        if line.strip() and not line.lstrip().startswith("#")
    }


def load_refresh_module():
    script = SKILL_ROOT / "scripts" / "refresh_dashboard.py"
    spec = importlib.util.spec_from_file_location("refresh_dashboard", script)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def all_input_slots():
    """Every declared read, including the three inputs whose sheets come from a spec.

    `orders`, `contract` and `signoff` list their sheets in the pipeline rather than in
    the registry, because the rest of the pipeline reads those specs for their business
    columns too. Expanding them here means a test asking "what does the pipeline read"
    gets the whole answer, not the part that happened to be written as a literal.
    """
    refresh_module = load_refresh_module()
    # Loading the pipeline puts its own directory on the import path, which is what makes
    # the sibling registry importable here by name.
    import sources

    slots = dict(sources.SLOTS)
    slots.update(sources.slots_for_families(
        refresh_module.ORDER_BOOK_SHEETS,
        refresh_module.PRICING_SHEETS,
        refresh_module.SIGNOFF_SHEETS,
    ))
    return slots


def test_approved_master_checksums_and_sheets():
    manifest = json.loads(
        (SKILL_ROOT / "config" / "master_manifest.json").read_text(encoding="utf-8")
    )
    for master in manifest["masters"]:
        path = SKILL_ROOT / master["path"]
        assert path.is_file(), f"Missing approved master: {master['path']}"
        digest = hashlib.sha256(path.read_bytes()).hexdigest()
        assert digest == master["sha256"]
        workbook = load_workbook(path, read_only=True, data_only=False)
        assert set(master["contains"]).issubset(workbook.sheetnames)


def test_pipeline_entrypoint_and_contract():
    config = json.loads(
        (SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8")
    )
    assert (SKILL_ROOT / config["entry_script"]).is_file()
    assert config["canonical_inputs"]["model"] == "rm_tracker_model.xlsx"
    assert config["canonical_inputs"]["zmat"] == "zmat.xlsx"
    assert config["coverage_days"] == {
        "critical_below": 15,
        "low_below": 30,
        "watch_below": 45,
        "target": 45,
    }
    assert config["boiler_material_groups"] == ["BOT", "COR", "AHT"]


def test_normalization_and_bucket_helpers():
    refresh = load_refresh_module()
    assert refresh.norm_od(22.2) == "22.23"
    assert refresh.norm_od(22.23) == "22.23"
    assert refresh.norm_surface("AP") == "AW"
    assert refresh.norm_surface("NR") == "AN"
    assert refresh.norm_thickness(1.5) == "1.6"
    assert refresh.norm_thickness(1.95) == "2"
    assert refresh.norm_thickness(3.4) == "3.5"
    assert refresh.make_ctl_bucket("42.7-0-4-ERW 1-FC", 3.2).endswith("-3.2")
    assert refresh.make_ctl_bucket("42.7-0-4-ERW 1-FC", 6.0).endswith("-1")


def test_bucket_strings_are_normalised_wherever_a_sheet_supplies_one():
    """A bucket is a join key, so it cannot depend on typing.

    The RM tracker's TVSM sheet wrote `25.4-0-2.5-ERW 1-FC ` on one row. That trailing
    space renders identically and is a different string, so the bucket became its own
    tracker row and its own pool: 60.304 MT of TVSM sales sat on a phantom row while the
    real bucket reported a 56 MT gap it had already dispatched.
    """
    refresh = load_refresh_module()
    assert refresh.norm_bucket("25.4-0-2.5-ERW 1-FC ") == "25.4-0-2.5-ERW 1-FC"
    assert refresh.norm_bucket("  25.4-0-2.5-ERW 1-FC") == "25.4-0-2.5-ERW 1-FC"
    assert refresh.norm_bucket("25.4-0-2.5-ERW  1-FC") == "25.4-0-2.5-ERW 1-FC"
    assert refresh.norm_bucket("25.4-0-2.5-ERW\xa01-FC") == "25.4-0-2.5-ERW 1-FC"
    assert refresh.norm_bucket("\xa0") is None
    assert refresh.norm_bucket("") is None
    assert refresh.norm_bucket(None) is None

    # Every sheet column that supplies a bucket must go through it.
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    for column in ('bucket_map["Bucket"]', 'bucket_map["CTL Bucket"]',
                   'schedule["Bucket"]', 'schedule["CTL Bucket"]',
                   'vsm["key"]', 'vsm_stock["key"]'):
        assert f"{column}.map(norm_bucket)" in script, (
            f"{column} reaches a join without norm_bucket"
        )


def test_a_sheets_own_grand_total_row_is_dropped_before_any_column_is_summed():
    """The RM tracker's TVSM sheet ends with its own grand total.

    It carries no key, so every per-bucket figure groups past it and stays right —
    which is why it survived. Only a whole-column sum sees it, and the sales-to-TVSM
    card read 1,981.672 MT where the truth is 990.836.
    """
    refresh = load_refresh_module()
    import pandas as pd

    frame = pd.DataFrame({
        "key": ["25.4-0-2-ERW 1-FC", "30-0-2-ERW 1-FC", "\xa0", None, "\xa0"],
        "VSM Sales": [10.0, 15.0, 1.5, None, 26.5],       # 26.5 == 10 + 15 + 1.5
        "VSM Stock": [4.0, 6.0, 0.0, None, 10.0],
    })
    totals = refresh.sheet_total_rows(frame, "key", ("VSM Sales", "VSM Stock"))
    assert list(totals) == [4], "the grand-total row is the one equal to all the others"
    kept = frame.drop(index=totals)
    assert kept["VSM Sales"].sum() == 26.5

    # A keyless row that is not a total stays, and a keyed row is never a candidate.
    assert 2 not in totals
    frame.loc[5] = ["40-0-2-ERW 1-FC", 26.5, 10.0]
    assert list(refresh.sheet_total_rows(frame, "key", ("VSM Sales",))) == []

    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert "vsm = vsm.drop(index=vsm_total_rows)" in script


def test_governed_constants_match_the_pipeline_contract():
    """The business rules live in two places; keep them from drifting apart."""
    refresh = load_refresh_module()
    config = json.loads(
        (SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8")
    )
    assert refresh.LONG_LENGTH_MIN_M == config["length"]["long_length_min_m"]
    assert refresh.HIGH_AGE_DAYS == config["stock_ageing"]["high_age_days"]
    assert refresh.RECEIVABLE_DUE_DAYS == config["receivables"]["due_days_from_invoice"]
    assert refresh.BILLING_DOC_TYPES == set(config["receivables"]["billing_doc_types"])
    assert refresh.OFFSET_NATURES == set(config["receivables"]["offset_natures"])
    assert refresh.CONVERSION_AGENT_OEM_BY_CODE == config["conversion_agent_oem_by_code"]
    assert refresh.TVS_PROXY_CUSTOMER_NAMES == set(config["tvs_proxy_customer_names"])
    assert refresh.TRANSIT_CUSTOMER_NAME == config["transit_customer_name"]
    assert refresh.STR_DESTINATION_PLANT == config["str_plan"]["destination_plant"]
    assert list(refresh.STR_SOURCE_PLANTS) == config["str_plan"]["source_plants"]
    assert refresh.STR_TARGET_DAYS == config["str_plan"]["target_days"]
    assert list(refresh.STR_CUSTOMERS) == config["str_plan"]["customers"]
    assert refresh.TRANSFER_INVOICE_MARKER == config["transfers"]["invoice_type_marker"]
    wip_to_fg = config["str_plan"]["wip_to_fg"]
    assert refresh.WIP_DESCRIPTION_PREFIX == wip_to_fg["wip_description_prefix"]
    assert refresh.FG_DESCRIPTION_PREFIX == wip_to_fg["fg_description_prefix"]
    assert refresh.FG_MATERIAL_TYPE == wip_to_fg["fg_material_type"]
    pricing = config["pricing"]
    assert list(refresh.PRICING_QUARTERS) == pricing["quarters"]
    assert refresh.PRICING_OPERATION_RATES == pricing["operation_rates_inr_per_ton"]
    assert refresh.PRICING_BORE_TOLERANCE_MM == pricing["bore_tolerance_mm"]
    assert {
        route: spec["sheet"] for route, spec in refresh.PRICING_SHEETS.items()
    } == pricing["sheets"]
    orders = config["order_book"]
    # Each sheet reports a different stage of the same commitment, so the column and
    # the header row are the two things that must never drift between code and config.
    assert {
        sheet: spec["quantity_column"]
        for sheet, spec in refresh.ORDER_BOOK_SHEETS.items()
    } == orders["open_quantity_column"]
    assert {
        sheet: spec["header"] + 1
        for sheet, spec in refresh.ORDER_BOOK_SHEETS.items()
    } == orders["header_row"]
    assert refresh.ORDER_PLANT_CODES == orders["plant_codes"]
    # A line marked in the remarks column is not live demand and must reach no pool.
    assert refresh.ORDER_EXCLUDE_REMARK == orders["exclude_remark"]
    assert {
        sheet: spec["remarks_column"]
        for sheet, spec in refresh.ORDER_BOOK_SHEETS.items()
    } == orders["remarks_column"]
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert 'orders_all[~orders_all["excluded"]]' in script, (
        "the pools must be built from the live book, not the whole sheet"
    )
    assert (
        refresh.VSM_LENGTH_MM_ABOVE_M
        == config["megh_length_bucketing"]["length_mm_above_m"]
    )
    # Each sheet states the order age differently — two name a day count, the third
    # only the date the STR was raised — so the basis is governed per sheet.
    assert {
        sheet: (
            f"{spec['age_column']} column" if spec["age_column"]
            else f"as-of date less {spec['age_date_column']}"
        )
        for sheet, spec in refresh.ORDER_BOOK_SHEETS.items()
    } == orders["age_source"]
    for spec in refresh.ORDER_BOOK_SHEETS.values():
        assert bool(spec["age_column"]) != bool(spec["age_date_column"]), (
            "a sheet states its age as a day count or as a date, never both nor neither"
        )


def test_sales_trend_reads_one_sheet_per_quarterly_extract():
    """The other sheets in those workbooks are working copies of Sheet1.

    `sales_q1.xlsx` carries Sheet5 and Sheet6 whose every row is already in Sheet1;
    reading them would double count those lines in the trend.
    """
    refresh = load_refresh_module()
    config = json.loads(
        (SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8")
    )
    trend = config["sales_trend"]
    assert refresh.SALES_TREND_SHEET == trend["sheet"]
    assert list(refresh.SALES_TREND_FILES) + ["sales.xlsx"] == trend["sources"]
    for name in refresh.SALES_TREND_FILES:
        assert name in config["optional_inputs"]
        assert config["required_sheets"][name] == [refresh.SALES_TREND_SHEET]
    assert refresh.MEGH_TVS_CUSTOMER_CODE in trend["segments"]["agent"]


def test_megh_only_plan_keys_are_not_read_as_governed_buckets():
    """The plan's key column carries two shapes with the same five parts.

    A governed bucket is `OD-ID-thickness-grade-endcondition`; a Megh-only size is
    `Megh-OD-ID-thickness-length`, where the last two parts mean something else
    entirely. Parsing one as the other builds a key out of "Megh" and reads a cut type
    off a length.
    """
    refresh = load_refresh_module()
    assert refresh.MEGH_KEY_PREFIX.upper() == "MEGH"
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    # Megh-only rows must never populate the governed-bucket column.
    assert 'vsm_stock["vsm_bucket"] = plan_key.where(~is_megh_only)' in script
    # And the plan's per-plant codes must be a fallback route to stock and sales, since
    # a size with no governed bucket can never be reached by the bucket join.
    assert "code_to_vsm_key" in script
    assert 'frame["material_key"].map(code_to_vsm_key)' in script


def test_the_megh_sku_key_is_the_bucket_with_its_length_appended():
    """The plan states the key; the pipeline no longer invents one.

    The old key was `OD-ID-thickness-length-grade-cuttype`, and its cut token came from
    the governed bucket's end condition. Where the plan and Bucketting disagree there —
    the plan keys 22.23 x 2.0 at 5.4 m as `…-ERW 1-FC` while Bucketting governs its own
    material code 2431251 to `…-ERW 1-PE` — the two sides built different keys, never
    met, and the tonnage left the tab for the unmapped queue.
    """
    refresh = load_refresh_module()
    assert refresh.bucket_vsm_key("22.23-0-2-ERW 1-PE", 5.951) == "22.23-0-2-ERW 1-PE-5.951"
    # A trailing space on the bucket must not fork the key, here as everywhere else.
    assert refresh.bucket_vsm_key("22.23-0-2-ERW 1-PE ", 5.951) == "22.23-0-2-ERW 1-PE-5.951"
    assert refresh.bucket_vsm_key(None, 5.951) is None
    assert refresh.bucket_vsm_key("22.23-0-2-ERW 1-PE", None) is None


def test_a_megh_sku_is_reached_only_by_a_statement():
    """A material code lands on a plan SKU because somebody said so, never by derivation.

    The owner's rule (19 Aug 2026): mapping is manual, so the data is deterministic. The
    two statements are the plan's own plant columns (`code_to_vsm_key`) and the owner's
    answer on the Megh queue (`sku_assignments`) — and nothing else. Until then the join
    *derived* a key from the governed bucket and the length, which had Bucketting, the
    TVSM ancillaries master, deciding what landed on the Megh tab: on the 14 Aug build,
    12 codes the plan itself names sat in the unmapped queue with 155.6 MT of 209.9
    because the two masters spell PE/FC or 6.001/6 differently, while 37.5 MT landed on
    SKUs no master had ever tied the code to.

    The derived key survives in exactly one place — the queue's suggestion column, where
    a person checks it instead of a join trusting it. So this test walks every block
    that writes a `vsm_key` onto a frame and asserts none of them calls
    `bucket_vsm_key`; the plan's own key construction (its stated `length key`, with
    bucket-plus-length only for a plan row stating none) and the queue's suggestion are
    the only callers left.
    """
    source = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")

    # Every frame-mapping block: from the assignment loop to the Megh customer filter.
    frames = source[source.index("for frame in (stock, sales, wip):")
                    :source.index('megh_codes = {"943209"')]
    assert "bucket_vsm_key" not in frames, (
        "a stock/sales/WIP frame derives its Megh SKU key again — the mapping is the "
        "plan's code columns and the owner's assignments, nothing inferred"
    )
    assert 'map(code_to_vsm_key)' in frames and "sku_assignments" in frames

    # The order-book join and the ledger history, which key the same tonnage elsewhere.
    orders = source[source.index("# Megh Steel's own orders, keyed to the VSM SKU")
                    :source.index("megh_keyed = ")]
    assert "bucket_vsm_key" not in orders
    history = source[source.index("megh_history = trend_all")
                     :source.index("megh_sales_months = ")]
    assert "bucket_vsm_key" not in history

    # And the suggestion is still offered: the queue is where the derived key now lives.
    queue = source[source.index('unmapped["suggested_key"]')
                   :source.index("megh_unmapped = [")]
    assert "bucket_vsm_key" in queue


def test_the_plans_length_key_is_corrected_only_where_it_could_not_join():
    """Both corrections leave the sheet as written and move only the derived key."""
    refresh = load_refresh_module()
    # A stray space renders identically and joins to nothing.
    assert refresh.norm_length_key("25.4-0-2.5-ERW 1-FC -5.95") == "25.4-0-2.5-ERW 1-FC-5.95"
    # Millimetres, where stock, sales and WIP all normalise to metres.
    assert refresh.norm_length_key(
        "59-30-1.6-ERW 1-FIN CUT-572.5") == "59-30-1.6-ERW 1-FIN CUT-0.5725"
    assert refresh.norm_length_key(
        "Megh-34.93-0-1.2-ERW 1-FC-0.620") == "Megh-34.93-0-1.2-ERW 1-FC-0.62"
    # Keys already in the join's own shape are left exactly alone.
    assert refresh.norm_length_key("12.7-0-1.2-ERW 1-PE-6") == "12.7-0-1.2-ERW 1-PE-6"
    assert refresh.norm_length_key("Megh-12.7-0-1.4-5.67") == "Megh-12.7-0-1.4-5.67"
    # A key stating no length keeps its own shape rather than acquiring an invented one.
    assert refresh.norm_length_key("50-50-1.6-ERW 1-PE") == "50-50-1.6-ERW 1-PE"
    # An empty cell is no key. `str(value or "")` would publish the string "nan" here,
    # because a float NaN is truthy.
    assert refresh.norm_length_key(float("nan")) is None
    assert refresh.norm_length_key("") is None
    assert refresh.norm_length_key(None) is None


def test_the_other_length_family_is_the_key_without_its_length():
    """Both sides of the join derive the family the same way — off the key itself."""
    refresh = load_refresh_module()
    assert refresh.key_family("22.23-0-2-ERW 1-PE-5.951") == "22.23-0-2-ERW 1-PE"
    # A grade carrying its own spaces and digits must not be mistaken for a length.
    assert refresh.key_family("38.1-0-1.6-ERW 2 MAHS-FC-6") == "38.1-0-1.6-ERW 2 MAHS-FC"
    # A Megh-only size has no governed bucket, and the family says so rather than
    # inventing one for other-length stock to be looked up under.
    assert refresh.key_family("Megh-12.7-0-1.4-5.67") == "Megh-12.7-0-1.4"


def test_the_megh_prefix_is_read_off_the_length_key():
    """The prefix marks a size going onward to RE or HMSIL rather than to TVSM.

    Four rows on the 18 September plan carry it on `length key` while their `key`
    column still names a governed TVS bucket (`34.93-0-1.2-ERW 1-FC`). The prefix is
    the statement of where the material goes, so it decides. No row prefixes `key`
    without also prefixing `length key`, so nothing is lost by reading the one.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert 'plan_length_key = vsm_stock["length key"].map(norm_length_key)' in script
    assert "plan_length_key.where(plan_length_key.notna(), plan_key)" in script


def test_the_vsm_stock_read_requires_the_plans_length_key():
    """A tracker without the column stops the run rather than resurrecting the guess."""
    script = SKILL_ROOT / "scripts" / "sources.py"
    spec = importlib.util.spec_from_file_location("sources", script)
    sources = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(sources)
    assert "length key" in sources.SLOTS["vsm_stock"].required


def test_bop_list_matches_the_config_and_claims_each_sku_once():
    """A listed line takes one SKU, and a SKU is claimed by one line.

    19.05 x 2.0 is on the list twice, at 5840 and 6000. Letting each line take its own
    nearest length hands 5.841 to both and leaves 5.8 unflagged; assigning nearest-first
    with one claim each gives them 5.841 and 5.8.
    """
    refresh = load_refresh_module()
    config = json.loads(
        (SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8")
    )
    bop = config["megh_bop"]
    assert len(refresh.MEGH_BOP_ITEMS) == bop["items"]
    assert sum(item[4] for item in refresh.MEGH_BOP_ITEMS) == bop["nos"]
    assert refresh.MEGH_BOP_LENGTH_TOLERANCE_MM == bop["length_tolerance_mm"]
    # The band is what stops a nominal length being read as a different size. 200 mm
    # pulled 19.05 x 2.0 x 6000 onto the plan's 5.8 m row; the owner's ruling is that
    # such a gap is a separate line item, so the tolerance stays inside a cut-length
    # rounding rather than a size difference.
    assert refresh.MEGH_BOP_LENGTH_TOLERANCE_MM <= 50
    # Every entry is (dim1, dim2, thickness, length_mm, nos, plant).
    for item in refresh.MEGH_BOP_ITEMS:
        assert len(item) == 6, item
        assert item[3] > 100, f"length is stated in millimetres: {item}"
        assert item[4] > 0 and isinstance(item[5], str)
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert "if index in claimed_lines or sku in claimed_skus:" in script, (
        "the assignment must claim each listed line and each SKU once"
    )
    # A listed size the plan has no row for becomes a row rather than a footnote.
    assert "megh_bop_added.append" in script
    assert '"megh_bop_added": megh_bop_added' in script


def test_dumps_folder_holds_the_masters_unchanged_and_is_never_deployed():
    """Two copies of a master that drift are worse than one.

    dumps/ makes the day's inputs self-sufficient, which means the three approved
    masters exist twice. They must stay byte-identical to the manifest. The folder must
    also stay out of the Vercel deployment: the GitHub repository is private, the
    deployment is not, and these files hold sales, receivables and contract prices.
    """
    repo = REPO_ROOT
    dumps = repo / "dumps"
    if not dumps.is_dir():
        return
    manifest = json.loads(
        (SKILL_ROOT / "config" / "master_manifest.json").read_text(encoding="utf-8")
    )
    for master in manifest["masters"]:
        copy = dumps / Path(master["path"]).name
        if copy.is_file():
            assert hashlib.sha256(copy.read_bytes()).hexdigest() == master["sha256"], (
                f"{copy.name} in dumps/ has drifted from the approved master"
            )
    assert "dumps/" in vercelignore_patterns(), (
        "dumps/ must not be served on the public deployment"
    )


def test_megh_length_bucketing_reads_the_plans_own_code_columns():
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    config = json.loads(
        (SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8")
    )
    columns = config["megh_length_bucketing"]["plant_code_columns"]
    used = re.search(r'plant_code_columns = \[c for c in \(([^)]*)\)', script)
    assert used, "the plan's per-plant code columns are no longer selected as a tuple"
    assert [c.strip().strip('"') for c in used.group(1).split(",") if c.strip()] == columns


def test_zmat_is_the_only_material_master_the_pipeline_reads():
    """A mapping extract is merged into zmat, never read beside it.

    Reading both means filtering and reasoning about two sources on every run to answer
    one question, and the filter is subtle enough to have been wrong once already.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert "material_mapping.xlsx" not in script
    config = json.loads(
        (SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8")
    )
    assert "material_mapping.xlsx" not in config["optional_inputs"]
    repo = REPO_ROOT
    assert not list((repo / "dumps").glob("material_mapping*")), (
        "the extract is merged into zmat and discarded, not stored"
    )


def test_material_mapping_merge_is_by_description_and_by_column_name():
    """Two traps, both of which produced wrong output before being caught.

    Filtering the extract on the material code rather than the description: it lists
    several codes per description, so a description zmat already resolved gains a second
    code, the description-to-code lookup sees the ambiguity and returns nothing, and 16
    stock-transfer lines lose the code they had.

    Aligning columns by position rather than by name: the two files hold the same 24
    columns in a different order, so renaming positionally transposes the file — it wrote
    descriptions into the code column and plants into the material type.
    """
    merge = (SKILL_ROOT / "scripts" / "merge_material_mapping.py").read_text(encoding="utf-8")
    assert "isin(known)" in merge and "normalise_description" in merge
    align = merge[merge.index("def align_columns"):merge.index("def main(")]
    assert "zip(" not in align, "columns must be matched by name, never zipped by position"
    assert "raise SystemExit" in align, "an unmatched column must stop the merge"


def test_order_code_source_tests_for_nan_explicitly():
    """A missing code arrives as NaN, which is truthy.

    Guarding on truthiness alone labels a blank cell as a code the sheet stated, and a
    line with no recovery at all as one zmat supplied — the same trap that built
    "nan-0.46" CTL buckets.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    block = script[script.index("def code_source_for("):script.index('frame["code_source"] = [')]
    assert "pd.notna(stated)" in block and "pd.notna(recovered)" in block


def test_transit_sheet_is_never_read():
    """PLANT STOCKS already holds the transit batches; reading both double counts."""
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert 'sheet_name="TRANSIT STOCK"' not in script
    config = json.loads(
        (SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8")
    )
    assert "TRANSIT STOCK" in config["do_not_read_sheets"]["stock.xlsx"]


def test_template_tab_row_wraps_so_the_page_cannot_overflow():
    """A non-wrapping tab row widens the document past the viewport."""
    template = (
        SKILL_ROOT / "assets" / "dashboard_template.html"
    ).read_text(encoding="utf-8")
    tabs_rule = next(line for line in template.splitlines() if line.strip().startswith(".tabs"))
    assert "flex-wrap: wrap" in tabs_rule


def test_documented_run_command_names_the_entry_script():
    skill_text = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
    assert "scripts/refresh_dashboard.py" in skill_text
    assert "--input-dir" in skill_text
    assert "--output-dir" in skill_text
    assert "--as-of" in skill_text


def test_config_lists_every_input_the_pipeline_reads():
    """Config drifts silently once it stops describing what the code does.

    `rfd_reconciliation` sat for a day describing a presence-only test the code no
    longer used, and `required_sheets` never gained the `vsm stock` sheet the Megh
    tracker depends on, because nothing compared the two.
    """
    config = json.loads(
        (SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8")
    )
    # Every read the pipeline makes is declared once, as a slot in `scripts/sources.py`.
    # Comparing the registry against the config beats grepping for `pd.read_excel` — it is
    # the same question asked of the thing that now answers it, and it keeps holding when a
    # second backend serves the same slots from somewhere other than a file.
    slots = all_input_slots()
    # A slot's first filename is the canonical one; any after it are older names still
    # accepted so that a workbook already archived under one keeps loading. The config
    # declares canonical names, so those are what it is compared against — and the
    # fallbacks are named here, because one added silently is a second file the pipeline
    # will read without anything saying so.
    read = {spec.files[0] for spec in slots.values()}
    declared = set(config["canonical_inputs"].values()) | set(config["optional_inputs"])
    assert read == declared, f"undeclared {read - declared}, stale {declared - read}"
    fallbacks = {name for spec in slots.values() for name in spec.files[1:]}
    assert fallbacks == {"july0626_rm_tracker_v1.xlsx", "rfd.xlsx"}


def test_every_slot_says_where_its_rows_are_kept():
    """A dump with nowhere to go is stored nowhere, and nothing says so.

    `raw_batches` takes any slot name, so a slot added to the registry without deciding
    where its rows are kept uploads cleanly, refreshes cleanly, and simply never reaches
    a table. The symptom arrives weeks later as a query returning less than it should.
    So the decision is required rather than defaulted: `table_for` raises on a slot that
    is in neither `TABLES` nor `SLOTS_WITHOUT_A_TABLE`, and this asks it of every slot
    the pipeline reads — the expanded `orders:`/`signoff:`/`contract:` families included,
    since those are built at run time and would otherwise be the ones missed.
    """
    load_refresh_module()
    import sources

    for slot in all_input_slots():
        # Raises rather than returning None where the slot is unaccounted for; None is
        # the answer for a slot that deliberately keeps nothing.
        sources.table_for(slot)

    # And the reasons stay honest: a slot listed as keeping nothing must still be a slot.
    known = set(all_input_slots()) | set(sources.SHEET_FAMILIES)
    unknown = set(sources.SLOTS_WITHOUT_A_TABLE) - known
    assert not unknown, f"SLOTS_WITHOUT_A_TABLE names slots that do not exist: {unknown}"


def snapshot_view_migration():
    return (REPO_ROOT / "supabase" / "migrations"
            / "20260814070000_a_view_per_snapshot_dump.sql").read_text(encoding="utf-8")


def test_the_snapshot_views_still_match_the_sheets_they_were_generated_from():
    """Two ways a generated view goes wrong, and the second one is silent.

    A slot with no view is a dump that cannot be read in SQL at all, which is loud enough
    to find. The one worth a test is positional drift: the stored grid is positional, so a
    column inserted in the *middle* of a sheet shifts every position after it, and the
    view goes on working perfectly while returning the wrong column under each name.
    Nothing errors and nothing is empty.

    It has happened once already. The `vsm stock` sheet gained a `length key` column at
    position 9 between one commit and the next, and `dump_vsm_stock` began labelling ten
    columns wrongly. A column added at the *end* is harmless by contrast — the view simply
    does not expose it — and this stays quiet about those, or it would fire on the
    ordinary event the whole design exists to absorb.
    """
    result = subprocess.run(
        [sys.executable, str(REPO_ROOT / "tools" / "generate_dump_views.py"), "--check"],
        capture_output=True, text=True,
    )
    assert result.returncode == 0, result.stderr


def test_every_snapshot_view_runs_as_the_invoker():
    """Without this the views hand raw receivables to anyone who can log in.

    A Postgres view runs as its *owner* by default, not its caller, so a view over
    `raw_rows` would return rows the base table's row-level security exists to withhold —
    sales lines, open invoices, customer rates. The whole access model is decided by one
    option per view, and there is nothing about a working query that would reveal it was
    missing.
    """
    written = snapshot_view_migration()
    created = re.findall(r"create or replace view public\.(dump_\w+)(.*)", written)
    assert created, "no views in the migration at all"
    for name, rest in created:
        assert "security_invoker = on" in rest, f"{name} runs as its owner"


def test_a_blank_row_is_recognised_in_one_place_and_not_thirteen():
    """This predicate has been wrong twice, and both times in every view at once.

    First written per column with `is null`, which never fires: an empty cell is stored as
    a JSON null, and `'[null]'::jsonb -> 0 is null` is false. Then rewritten as one
    expression and negated, which kept only the blank rows — every view returned padding
    and nothing else. Each was a one-character fix in thirteen places.
    """
    written = snapshot_view_migration()
    assert "create or replace function public.dump_row_has_data" in written
    for view_body in written.split("create or replace view")[1:]:
        assert "public.dump_row_has_data(r.row)" in view_body, (
            "a view filtering blank rows by hand will be the one left behind"
        )
    assert "jsonb_array_elements" not in written.split("-- ---- The views")[-1], (
        "the row-emptiness test belongs in the function, not inlined in a view"
    )


def test_the_pipeline_reads_nothing_except_through_the_source_registry():
    """One inline `pd.read_excel` is enough to break a second backend, silently.

    The whole point of the registry is that a backend other than the dumps folder can
    serve every frame the pipeline needs. A read that goes straight to a file bypasses
    it: run against Postgres, that one frame would still come off disk — stale, or
    absent, and in neither case announced. The failure would surface as a wrong number
    on a page, not as an error.

    So the pipeline gets its frames from `src` and from nowhere else. The registry
    itself is exempt, since performing the reads is its job.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    for forbidden in ("pd.read_excel", "pd.ExcelFile", "pd.read_csv"):
        assert forbidden not in script, f"{forbidden} belongs in scripts/sources.py"
    # Reaching into the input directory by hand is the same bypass wearing a hat.
    assert "input_dir /" not in script

    registry = (SKILL_ROOT / "scripts" / "sources.py").read_text(encoding="utf-8")
    assert "pd.read_excel" in registry


def test_config_lists_every_sheet_the_pipeline_reads():
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    config = json.loads(
        (SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8")
    )
    read = set(re.findall(r'sheet_name="([^"]+)"', script))
    declared = {s for sheets in config["required_sheets"].values() for s in sheets}
    assert read <= declared, f"sheets read but not declared: {read - declared}"


def test_account_management_round_trip(tmp_path):
    """Accounts live in access.json, so creating one must not need a rebuild."""
    access = tmp_path / "access.json"
    access.write_text(json.dumps({
        "admin": "apoorv",
        "salt": "test-salt",
        "accounts": {"apoorv": "x"},
        "tabs": [{"view": "meghView", "label": "Megh"},
                 {"view": "llView", "label": "LL"}],
        "visible": {},
    }), encoding="utf-8")
    script = SKILL_ROOT / "scripts" / "manage_users.py"

    def run(*args):
        return subprocess.run(
            [sys.executable, str(script), "--access", str(access), *args],
            capture_output=True, text=True,
        )

    assert run("add", "someone", "--password", "pw", "--tabs", "meghView").returncode == 0
    saved = json.loads(access.read_text(encoding="utf-8"))
    expected = hashlib.sha256(b"someone:test-salt:pw").hexdigest()
    assert saved["accounts"]["someone"] == expected
    assert saved["visible"]["someone"] == ["meghView"]
    # The grant follows the tab order rather than the argument order.
    run("grant", "someone", "--tabs", "llView", "meghView")
    assert json.loads(access.read_text(encoding="utf-8"))["visible"]["someone"] == [
        "meghView", "llView"
    ]
    # Guards: an unknown tab, and the admin account.
    assert run("grant", "someone", "--tabs", "nope").returncode != 0
    assert run("remove", "apoorv").returncode != 0
    assert run("remove", "someone").returncode == 0
    saved = json.loads(access.read_text(encoding="utf-8"))
    assert "someone" not in saved["accounts"] and "someone" not in saved["visible"]


def test_customer_sku_history_is_keyed_on_the_customers_own_codes():
    """The tracker's history must not be joined through `display_by_code`.

    That map exists for the STR plan and drops any SAP code used under more than one
    Helper Customer, which is right when the decision moves stock and wrong when it
    only displays history. Three codes are shared on the 31 July build, and routing
    through it left Balaji Press Product and ELKAYEM AUTO Hosur with an empty table
    across 45 tracker lines — indistinguishable from having bought nothing.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    # The code/customer maps are built once, above the trend tables that first need them.
    assert "codes_by_display.setdefault(display, set()).add(code)" in script
    assert "displays_by_code.setdefault(code, set()).add(display)" in script
    history = script[script.index("history_keyed = keyed.copy()"):]
    history = history[:history.index("customer_sku_history.sort")]
    assert 'history_keyed["customer_key"].isin(codes)' in history
    # `display_by_code` is the STR bridge and must not appear here. Matched on a word
    # boundary, because `displays_by_code` and `unique_display_by_code` are different
    # maps built for the trend grouping and are not what this rule is about.
    assert not re.search(r"(?<![_a-zA-Z])display_by_code", history), (
        "the history join must read the customer's own codes, not the STR bridge"
    )
    # An empty table has two causes and states neither, so both must reach the page.
    assert '"customer_history_notes": customer_history_notes' in script
    assert '"shared_codes_with"' in script and '"reason"' in script

    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    assert "customer_history_notes" in template, (
        "the page must say why a history is empty rather than showing a blank table"
    )
    # Per-row totals repeat when one SKU is scheduled at two plants, so the column
    # must not be summable — a total across it would double count.
    row = template[template.index("const customerRowHtml"):]
    row = row[:row.index("</tr>`;")]
    assert "history_detail_key" in row
    assert 'data-sum' not in row.split("history_detail_key")[1]


def test_the_package_sits_where_claude_code_discovers_it():
    """Layout check, because two of these paths are load-bearing and fail quietly.

    The skill must sit at `.claude/skills/<name>/SKILL.md` to be discovered, and the
    repository root must be reachable from it — the dumps/ and .vercelignore checks
    below derive `repo` from this file's depth, and when the package moved under
    `.claude/` they would have gone on passing while looking in the wrong directory.
    """
    assert SKILL_ROOT.name == "refresh-tvsm-dashboard"
    assert SKILL_ROOT.parent.name == "skills"
    assert SKILL_ROOT.parent.parent.name == ".claude"
    assert (SKILL_ROOT / "SKILL.md").is_file()
    # Frontmatter is what makes it a skill rather than a markdown file.
    head = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8").split("---")[1]
    assert "name: refresh-tvsm-dashboard" in head
    assert "description:" in head

    # The repository root, proved rather than counted.
    assert (REPO_ROOT / ".vercelignore").is_file(), "REPO_ROOT is not the repository root"
    assert (REPO_ROOT / "index.html").is_file()
    assert (REPO_ROOT / ".claude").is_dir()

    # Everything under .claude/ is commercial or internal, and the deployment is public.
    ignored = vercelignore_patterns()
    assert ".claude/" in ignored, (
        "the skill package holds the approved masters, including contract prices"
    )
    for stale in ("skill/", "context/", "CLAUDE.md"):
        assert stale not in ignored, f"{stale} no longer exists; drop it from .vercelignore"


def test_a_sales_line_held_by_two_extracts_is_counted_once():
    """A line in two sources must not be counted twice, and must not be lost.

    Both halves failed on the same day. The daily dump only covers the month in
    progress, so on 1 August July existed in no source at all and 3,344.836 MT — the
    most recent complete month — dropped out of a six-month view. Archiving the closed
    month's dump fixes that only if the overlap it creates is handled: while the daily
    dump still covers a month, the archive of that month is sitting beside it.

    This used to be answered by letting each source claim the months no earlier source
    had. That deduplicated whole months because there was no finer key to hand, and it
    made a *partial* backfill impossible: an extract that finally carried Jamshedpur for
    a closed month could only replace that month or be ignored, never merge into it. The
    ledger keys on the billing line instead, so this is now a property of the key — and
    is tested as behaviour rather than as the shape of the code.
    """
    load_refresh_module()
    import pandas as pd
    import sources

    def extract(*docs):
        return pd.DataFrame([
            {"Billing  Document Number": doc, "Billing Item": item,
             "BILLING  DATE": date, "Quantity": qty}
            for doc, item, date, qty in docs
        ])

    class TwoExtracts(sources.Sources):
        def available(self, slot):
            return slot in ("sales", "sales_jul")

        def frame(self, slot, *, sheet=None):
            # The same invoice line in both, as a re-sent day really arrives: the daily
            # dump still covers the month whose archive is sitting beside it. Plus one
            # line only the archive has, which is what a backfill looks like.
            if slot == "sales":
                return extract((4731002954.0, 10.0, "2026-07-31", 1000.0))
            return extract((4731002954, 10, "2026-07-31", 1000.0),
                           (4731002955, 20, "2026-07-30", 500.0))

    ledger = TwoExtracts().sales_ledger()
    assert len(ledger) == 2, "the shared line must appear once and the archive's line must survive"
    assert float(ledger["Quantity"].sum()) == 1500.0, "no line counted twice"
    # Float out of one extract and int out of the other is the normal case, not an edge
    # one: a daily dump carries a grand-total row and so reads float, an archive has none
    # and reads int. Untexted, the same invoice keys two ways and is stored twice.
    assert set(ledger["billing_document"]) == {"4731002954", "4731002955"}


def test_the_sales_grand_total_row_never_enters_the_ledger():
    """Every daily dump ends with the sheet's own total, and it must not be a line.

    No customer, no date, no material, no billing document — and `Quantity` holding the
    sum of the column, 968,438 kg against a month's real 786 lines on the 7 August file.
    A ledger that deduplicated on a blank key would keep one of them.
    """
    load_refresh_module()
    import pandas as pd
    import sources

    frame = pd.DataFrame([
        {"Billing  Document Number": 4731002954.0, "Billing Item": 10.0,
         "BILLING  DATE": "2026-08-01", "Quantity": 1000.0},
        {"Billing  Document Number": None, "Billing Item": None,
         "BILLING  DATE": None, "Quantity": 968438.483},
    ])
    kept = sources.sales_line_keys(frame)
    assert len(kept) == 1
    assert float(kept["Quantity"].sum()) == 1000.0


def test_customer_code_oem_reads_the_whole_sales_history():
    """A code's OEM does not depend on where in the month the refresh runs.

    `code_oem` was built from the daily dump alone. On 3 August that dump held three
    days and 125 lines, most schedule customers had not bought yet, and schedule OEM
    resolution fell to 0.6675 against a 0.99 publication gate — a FAIL produced by the
    calendar rather than by the data.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    block = script[script.index("    code_oem = ("):]
    block = block[:block.index(".to_dict()")]
    # `sales_all` is every line the ledger holds; `sales` is the published month alone.
    # Reading the month here is the defect this test exists for, so it is the one that
    # must not appear.
    assert "sales_all" in block, (
        "code_oem must read the whole ledger, not only the month in progress"
    )
    assert not re.search(r"\bsales\b(?!_all)", block), (
        "code_oem must not narrow to the published month"
    )
    # The gate this protects.
    assert 'failures.append("Schedule OEM resolution is below 99%.")' in script


def test_a_schedule_from_another_month_is_announced_on_the_page():
    """Demand and dispatch must be the same month's, or every balance is nonsense.

    The dumps roll into a new month before the next schedule workbook arrives, so the
    page compares last month's demand with this month's sales. On 3 August that read as
    4,200.991 MT open against 971.688 MT the week before — the calendar, not a collapse
    in dispatch. It has to say so rather than leave a reader to work it out.
    """
    refresh = load_refresh_module()
    # The sheet is named for the month it covers and the owner renames it as the month
    # rolls, so it is resolved from the as-of date rather than declared. Two earlier
    # versions of this test hardcoded a sheet name and a month number, and each broke on
    # the next workbook — first when August demand arrived on a sheet still called
    # "Schedule July", then when the sheet was renamed "Schedule August".
    assert refresh.SCHEDULE_SHEET_PREFIX == "Schedule "
    assert refresh.MONTH_NAMES[7] == "August"
    assert len(refresh.MONTH_NAMES) == 12
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert 'wanted = f"{SCHEDULE_SHEET_PREFIX}{as_of_month_name}"' in script
    # The row bound must not be a count: the sheet grew from 742 rows to 748 between two
    # workbooks, and a fixed slice would have dropped the six new lines in silence.
    assert ".iloc[:742]" not in script
    assert 'schedule["Helper Customer"].notna()' in script
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert "if schedule_month is not None and as_of_date.month != schedule_month:" in script
    # A warning, not a failure: the rest of the page — stock, ageing, receivables,
    # the Megh plan — is current and worth publishing.
    guard = script[script.index(
        "if schedule_month is not None and as_of_date.month != schedule_month:"):]
    guard = guard[:guard.index("if qc[")]
    assert "warnings.append" in guard and "failures.append" not in guard
    assert 'src.frame("schedule", sheet=schedule_sheet)' in script, \
        "the sheet must be the resolved one"


def test_near_gauge_thicknesses_group_onto_a_governed_one():
    """A wall a customer writes must land on a thickness Bucketting actually governs.

    Bucketting holds 256 buckets at 1.20 and none at 1.21 or 1.22, so both are ways of
    writing the same gauge. 1.22 was missing from the table, and Srikam, Rajsriya and
    NMPL lines written that way met no bucket at all — 85,500 pieces of the August
    schedule reaching no row.
    """
    refresh = load_refresh_module()
    for written, governed in ((1.21, "1.2"), (1.22, "1.2"), (1.63, "1.6"),
                              (2.03, "2"), (2.34, "2.3"), (2.75, "2.8")):
        assert refresh.norm_thickness(written) == governed, written
    # A real gauge of its own must not be swallowed by the grouping.
    assert refresh.norm_thickness(2.95) == "2.95"


def test_near_value_dimensions_fold_onto_the_governed_one():
    """Customers write diameters and walls Bucketting does not carry.

    Each of these is a value Bucketting holds under one number and nothing near: 22.23
    (never 22.2), 41.28 (never 41.3), 1.2 (never 1.21 or 1.22), 1.6 (never 1.62 or
    1.63). Folding recovers the size; not folding sends the demand to no bucket at all.
    """
    refresh = load_refresh_module()
    for written, governed in ((22.2, "22.23"), (41.3, "41.28"), (41.28, "41.28")):
        assert refresh.norm_od(written) == governed, written
    for written, governed in ((1.21, "1.2"), (1.22, "1.2"), (1.62, "1.6"), (1.63, "1.6")):
        assert refresh.norm_thickness(written) == governed, written
    # A diameter with no entry passes through rather than snapping to a neighbour.
    assert refresh.norm_od(25.4) == "25.4"
    assert refresh.norm_od(41.0) == "41"


def test_schedule_tonnage_never_comes_from_bucketting_kg_per_nos():
    """That column carries two different units and cannot be read as one.

    On a cut-length row `kG/nos` is kilograms per piece; on a 5.6 m or 6 m row it is
    kilograms per metre. 25.4 x 3.5 shows 1.889346 for both a 0.247 m piece weighing
    0.4667 kg and the 5.6 m row, because the second is the per-metre figure. Reading it
    as per-piece gave a 935 mm piece a 283 mm piece's weight; reading it as per-metre
    left Narasipur 25 MT light. The pipeline's own formula agrees with the trustworthy
    rows to 0.05% and governs every other tonnage on the page.
    """
    builder = (SKILL_ROOT / "scripts" / "schedules" / "build_supplement.py").read_text(
        encoding="utf-8")
    assert "kG/nos" in builder, "the reasoning for not using it must stay recorded"
    body = builder[builder.index("def tonnage("):]
    body = body[:body.index("\nfor _, line in matched")]
    assert "by_dims" not in body, "tonnage must not read a weight out of Bucketting"
    assert "stated" in body


def test_dumps_shows_the_current_month_and_archives_the_rest():
    """`dumps/` answers what this month is built from, not what was ever sent.

    Left alone it answers the second question, because each day overwrites the file
    before it and the month boundary passes unmarked.
    """
    archiver = SKILL_ROOT / "scripts" / "archive_month.py"
    assert archiver.is_file()
    text = archiver.read_text(encoding="utf-8")
    # A closed month's sales dump is archived by being renamed into a trend source and
    # is still read daily; moving it would drop that month out of the trend.
    assert r"^sales_[a-z]{3}\.xlsx$" in text
    assert r"^sales_q\d\.xlsx$" in text
    assert '"zmat.xlsx", "contract.xlsx"' in text
    # The schedule workbook belongs to its month and must not be pinned to the top.
    keep = text[text.index("KEEP_AT_TOP = {"):]
    keep = keep[:keep.index("}")]
    assert "rm_tracker_model" not in keep
    assert "shutil.move" in text and "shutil.copy" not in text, "archiving moves, never copies"

    config = json.loads((SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8"))
    assert config["canonical_inputs"]["model"] == "rm_tracker_model.xlsx"
    # The slot accepts the month-neutral name first and the old one after it, so a
    # workbook already archived under the July name still loads.
    model_files = all_input_slots()["schedule"].files
    assert model_files[0] == "rm_tracker_model.xlsx"
    assert "july0626_rm_tracker_v1.xlsx" in model_files, \
        "the old name stays readable as a fallback"

    repo = REPO_ROOT
    if (repo / "dumps").is_dir():
        months = [p.name for p in (repo / "dumps").iterdir()
                  if p.is_dir() and re.fullmatch(r"\d{4}-\d{2}", p.name)]
        for month in months:
            # A canonical name appearing in both places is the point of the layout —
            # dumps/stock.xlsx is this month's, dumps/2026-07/stock.xlsx is July's. What
            # must never happen is the same *bytes* in both, which would mean a month was
            # copied rather than moved and the two can drift apart.
            for archived in (repo / "dumps" / month).iterdir():
                current = repo / "dumps" / archived.name
                if current.is_file() and archived.is_file():
                    assert current.read_bytes() != archived.read_bytes(), (
                        f"{archived.name} is byte-identical in dumps/ and dumps/{month}/ "
                        "— archiving moves, it does not copy"
                    )


def test_order_signoff_splits_each_plant_on_its_own_terms():
    """The three sheets agree on nothing except that a material code names the line.

    jsr marks sign-off with a Y/N flag, so a line's whole balance falls to one side.
    Hosur states the signed tonnage against the order and the rest is unsigned. Khopoli
    states both halves outright. Read as one convention, two of the three come out wrong.
    """
    refresh = load_refresh_module()
    config = json.loads((SKILL_ROOT / "config" / "pipeline.json").read_text(encoding="utf-8"))
    assert refresh.SIGNOFF_FILE == "signoff.xlsx"
    assert set(refresh.SIGNOFF_SHEETS) == {"jsr", "hosur", "khopoli"}
    assert refresh.SIGNOFF_FILE in config["optional_inputs"]

    jsr = refresh.SIGNOFF_SHEETS["jsr"]
    assert jsr["flag"] == "Sign Off" and jsr["signed_when"] == "Y"
    # `Sign Off Qty` exceeds `Bal to Desp` on 9 of 27 rows, so it is measuring something
    # else and cannot be one part of that total.
    assert "Sign Off Qty" not in jsr.values()
    assert refresh.SIGNOFF_SHEETS["khopoli"]["unsigned"] == "Non Sign Off"
    assert refresh.SIGNOFF_SHEETS["hosur"]["unsigned"] is None, "hosur's remainder is derived"

    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    # A sheet signing off more than it ordered must not lend the difference back.
    assert "(quantity - signed).clip(lower=0)" in script
    # Both trackers carry both halves, and a code reaching neither is named rather than
    # dropped: its tonnage is in the file and on no row.
    for column in ('ll_tracker["signoff_mt"]', 'll_tracker["non_signoff_mt"]'):
        assert column in script
    assert 'row["signoff_mt"]' in script and 'row["non_signoff_mt"]' in script
    assert '"Order sign-off", SIGNOFF_FILE' in script

    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    assert template.count("signoffCells(r,") == 2, "one call on each of the two trackers"
    assert "<option>Order sign-off</option>" in template, (
        "a queue with no way to filter to it is hard to find"
    )
    # Every table's totals row must match its header, or refreshAllTotals writes past the
    # end of the row and rendering stops — which is what adding these columns did first.
    import re
    th = re.compile(r"<th[\s>]")
    for table in re.findall(r'<table id="(\w+)">', template):
        body = template[template.index(f'<table id="{table}">'):]
        if "</thead>" not in body:
            continue
        body = body[:body.index("</thead>")]
        if '<tr class="totals-row">' not in body:
            continue
        split = body.index('<tr class="totals-row">')
        assert len(th.findall(body[:split])) == len(th.findall(body[split:])), (
            f"{table}: totals row does not match its header"
        )


def test_each_drill_down_quantity_column_totals_its_own_field():
    """A breakup with two quantity columns must not total one of them three times.

    Every detail schema had exactly one quantity column until sign-off arrived with
    three — signed, not signed, order — and the modal's totals row summed `qty` once and
    wrote that figure under every heading. The 22.23-0-2-ERW 1-PE breakup then read
    337.71 MT signed, 337.71 MT not signed and 337.71 MT ordered: the same number three
    times, saying both that everything was signed off and that nothing was.
    """
    refresh = load_refresh_module()
    fields = [c["field"] for c in refresh.SIGNOFF_DETAIL_COLUMNS if c.get("kind") == "qty"]
    assert len(fields) == len(set(fields)) > 1, (
        "the sign-off breakup is the schema that makes a shared total wrong"
    )

    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    assert "rows.reduce((sum, row) => sum + Number(row[c.field] || 0), 0)" in template, (
        "each quantity column totals the field it displays"
    )
    assert "rows.reduce((sum, row) => sum + Number(row.qty || 0), 0)" not in template, (
        "one total shared across every quantity column is the defect"
    )


def test_a_per_row_average_is_never_added_down_its_own_column():
    """The customer history's last two columns describe a rate, not a quantity.

    Every column of that table was totalled the same way, so "Avg active month MT" — a
    per-SKU figure — was added down its column and read 214.291 MT for a customer whose
    busiest month was 260 and whose quiet ones were 129, because a SKU selling in three
    months and one selling in eight each contributed a single number. "Months" added the
    same way reached 197 over an eight-month window. Both now come off the month columns
    beside them.
    """
    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    assert 'data-month="1"' in template, "the month columns have to be identifiable"
    assert '{ label: "Months", digits: 0, agg: "months" }' in template
    assert '{ label: "Avg active month MT", agg: "avg", over: "Total MT",' in template
    # Divide the unrounded column total, not the months added up: data.json rounds each
    # month and the total separately, so the two disagree in the last place and a totals
    # row that does not check out against the cells beside it invites doubt.
    assert 'header.dataset.aggOver' in template
    # The denominator is the months actually sold in, not the width of the window: a
    # size dropped in April must not read as though it still sells every month at a
    # quarter of the rate.
    assert "const active = monthColumns.filter(index => (sums[index] || 0) > 0);" in template
    assert "const average = tonnage / active.length;" in template
    assert "(sums[over] || 0)" in template
    # The label has to say the row is no longer one kind of figure throughout.
    assert 'trailing.some(h => h.agg) ? "Total · avg" : "Total"' in template


def test_cut_length_history_carries_pieces_beside_the_tonnage():
    """A cut length is ordered in pieces, so a CTL history in tonnage alone is misread.

    12.197 MT of a 2.59 m piece is a figure nobody schedules against; 4,710 pieces is the
    one the customer's own schedule states. Both frames that carry a `length_type` now
    emit the months in pieces as well, and the page shows them on CTL rows only — a long
    length is bought by weight and a piece count there says nothing a buyer acts on.
    """
    refresh = load_refresh_module()
    labels = [c["label"] for c in refresh.SKU_HISTORY_DETAIL_COLUMNS]
    assert "Pieces" in labels and "Sales" in labels, "the drill-down carries both units"

    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    # Pieces come off the same frame as the tonnage, so the two can never disagree.
    # Bucket trend, customer SKU trend, and the tracker's own history: every frame that
    # runs months across its columns emits them in pieces too.
    assert script.count('value="sales_nos", digits=0') == 3
    for field in ('"months_nos": months_nos', '"avg_active_month_nos"'):
        assert field in script

    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    assert 'r.length_type === "CTL" && Number(nos)' in template, (
        "pieces belong on cut-length rows, not on long lengths"
    )
    # One renderer for both tables: the customer tracker's history and the past sales
    # trend must not drift into showing the same month two different ways.
    assert template.count("months.map(m => monthCell(r, m") == 2
    # Two figures in one cell read back as one run of text — "12.197 MT4,710 nos" parses
    # as neither — so such a cell states what it pastes as.
    assert "if (cell.dataset.copy !== undefined) return cell.dataset.copy;" in template
    assert 'data-copy="${copyAs === undefined ? "" : copyAs}"' in template
    assert 'totalCell.dataset.copy = fmt(sum, Number(digits)).replace(/,/g, "")' in template, (
        "the clipboard gets the displayed figure, not the running float behind it"
    )


def test_long_length_tracker_carries_the_last_complete_months_billing():
    """The tracker reads one month against one month's stock, which cannot say if it is normal.

    The bucket's billing history already exists for the trend tab, so the tracker shows
    the last complete month beside the current one and opens the window on a click. Last
    *complete* month, not the last month held: the window runs to the month in progress,
    and five days of August set against a full July reads as a collapse.
    """
    refresh = load_refresh_module()
    labels = [c["label"] for c in refresh.LL_HISTORY_DETAIL_COLUMNS]
    assert labels[0] == "Month", "the card is read down the months"
    # Both parties are named rather than merged: a month where Megh replaced an ancillary
    # on the same size is not a flat month.
    assert "TVSM ancillaries" in labels and "Megh Steel 943209" in labels

    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert "complete_months = [m for m in trend_months if m < as_of_month]" in script
    assert '"LLHISTORY": LL_HISTORY_DETAIL_COLUMNS' in script
    # A bucket that has never been billed has nothing to open, and a button onto an empty
    # card is worse than a plain figure.
    assert 'history_key_by_bucket[bucket] = f"LLHISTORY|{bucket}" if months else None' in script

    # The frame is built once. Order book and sign-off are added to `ll_tracker` as
    # columns long after `ll_rows` was last touched, so rebuilding it from that list
    # drops them — which it did, publishing 0 MT ordered and 0 MT signed off across all
    # 90 rows.
    assert script.count("ll_tracker = pd.DataFrame(ll_rows)") == 1
    assert 'll_tracker["last_month_sales_mt"] = ll_tracker["bucket"].map(' in script
    # `kind: "mt"` is not a licence to add: it is also weight per metre and rupees per
    # metre. A column that totals says so.
    assert all(c.get("add") for c in refresh.LL_HISTORY_DETAIL_COLUMNS
               if c["field"] in ("direct_mt", "megh_mt"))

    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    assert 'const summable = c => c.kind === "qty" || c.add === true;' in template
    # The header names the month. Two adjacent columns holding different months, both
    # labelled only by position, invite exactly one mistake.
    assert 'id="llLastMonthHead"' in template
    assert "`${monthLabel(lastFull)} sales`" in template
    assert "r.last_month_sales_mt" in template


def test_a_month_by_month_card_closes_on_an_average_month():
    """Adding a history gives the size of the window, not the size of a month.

    The long-length tracker's billing card and the customer tracker's SKU history card
    both list one row per month. Their bottom row totalled, so a bucket billed steadily
    at ~130 MT a month closed on 928.982 MT — the eight-month window, a figure no
    schedule is read against. Both now close on an average month, over the months that
    moved, which is the denominator the tracker's own average column already uses.
    """
    refresh = load_refresh_module()
    # The pieces column has to average with the tonnage beside it, and `kind: "num"`
    # alone does not make a column summable.
    assert any(c["field"] == "nos" and c.get("add") for c in refresh.SKU_HISTORY_DETAIL_COLUMNS)

    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    assert 'const AVERAGE_BY_MONTH = new Set(["LLHISTORY", "SKUHISTORY"]);' in template
    assert 'const over = byMonth ? rows.length : 1;' in template
    # Labelled, so an averaged row is never read as a total.
    assert 'byMonth ? "Average month" : "Total"' in template
    # Every other card still totals: a stock pool or an order book is a quantity, and
    # dividing it by its line count would mean nothing.
    assert 'fmt(total / over, places)' in template
    # A pieces column beside a tonnage is a count: no unit suffix, no decimals.
    assert 'const counts = c.kind === "num";' in template


def test_the_trackers_history_cell_is_an_average_month():
    """The figures beside it are one month's schedule and one month's dispatch.

    The cell held the SKU's total over the whole trend window, so a size running at 12 MT
    a month showed 86.641 against a 12 MT schedule and read as running six times its real
    rate. It is now the average month, over the months that moved — the same rule the
    history table's own column and the drill-down card use, so all three agree.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    assert '"history_total_mt"' not in script, "the window total is not what the line needs"
    assert 'schedule_export["history_avg_month_mt"]' in script
    # How many months it averages over decides whether the figure means anything.
    assert 'schedule_export["history_months_active"]' in script

    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    # Both customer tables carry it, and the heading has to say it is an average or a
    # figure beside a single month's schedule reads as another single month.
    assert template.count('<th class="num">Avg month sales</th>') == 2
    assert "r.history_avg_month_mt" in template
    assert "Averaged over the ${r.history_months_active} month" in template


def test_the_past_sales_trend_switches_between_tonnes_and_pieces():
    """A cut length is ordered in pieces and a long length by weight.

    A trend held only in tonnage answers half the question, so every frame on the tab
    carries both and one switch reads the whole tab in the chosen unit — headline cards,
    month strip, bucket table, customer SKU table and plant summary together.
    """
    refresh = load_refresh_module()
    labels = [c["label"] for c in refresh.TREND_DETAIL_COLUMNS]
    # The breakup cards carry both units whatever the tab is set to, so switching never
    # changes what a breakup says.
    assert "Pieces" in labels and "Sales" in labels

    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    for field in ('"months_nos": months_nos', '"total_nos": round(sum(months_nos.values()), 0)',
                  '"direct_nos"', '"megh_nos"', '"segment_totals_nos"', '"total_nos": round(float('):
        assert field in script, field

    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    assert '<select id="trendUnit">' in template
    # One switch for the tab, not one per table: two that can disagree is worse than none.
    assert "trendPlantUnit" not in template
    assert template.count('const pieces = inPieces_();') == 2
    # Month columns take their precision from the unit — pieces are whole numbers.
    assert "function trendHead(id, leading, trailing, monthDigits = 3)" in template
    assert 'data-sum="${monthDigits}"' in template


def test_a_missing_access_file_says_so_rather_than_blaming_the_password():
    """The admin fallback exists so a missing file cannot lock everyone out.

    Its side effect is that every non-admin account then fails the digest check and gets
    "Username or password is not recognised" — sending the holder to re-check a password
    that was never wrong. The page names the real cause instead.
    """
    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    assert 'if (!accessPublished && username !== (access.admin || "apoorv"))' in template
    assert "access.json did not load, so only the admin account can sign in here." in template
    # The admin fallback itself stays: one account able to get in and see what is wrong
    # beats everyone locked out.
    assert "const FALLBACK_ACCOUNTS = {" in template


def test_the_web_apps_drill_down_carries_the_same_totals_rules_as_the_page():
    """The breakup panel is ported; the three rules that govern its footer come with it.

    Each of them was wrong on the static page once, and none of them is expressible in a
    type: a formula breakup that gets totalled, a history that gets added instead of
    averaged, and three quantity columns that all report the same field are all valid
    TypeScript. So they are asserted here against the component that now renders them.
    """
    panel = (REPO_ROOT / "app" / "dashboard" / "[view]" / "detail.tsx").read_text(encoding="utf-8")

    # A formula breakup lists its inputs and its result; a column of those adds to nothing.
    assert 'const NO_TOTAL = new Set(["LLCOVERAGE", "LLGAP45", "LLGAP", "BALANCE"]);' in panel
    # A breakup whose rows are months closes on an average month, and says so.
    assert 'const AVERAGE_BY_MONTH = new Set(["LLHISTORY", "SKUHISTORY"]);' in panel
    assert "const over = byMonth ? rows.length : 1;" in panel
    assert 'byMonth ? "Average month" : "Total"' in panel
    # Each quantity column totals the field it displays. One total computed from `qty`
    # and repeated across the sign-off split read as a bucket both fully signed off and
    # fully outstanding.
    assert "rows.reduce((sum, row) => sum + (Number(row[c.field]) || 0), 0)" in panel
    assert "row.qty" not in panel, "a shared total across quantity columns is the defect"
    # `kind: "mt"` is not summable — it is also weight per metre and rupees per metre.
    assert 'c.kind === "qty" || c.add === true' in panel
    # A pieces column beside a tonnage is a count: no unit suffix and no decimals.
    assert 'const counts = c.kind === "num";' in panel


def test_the_web_app_fetches_the_drill_down_layouts_on_every_tab():
    """Without them every breakup falls back to the source-and-quantity layout.

    `detail_columns` is a scalar on every build and is not admin-only, so it rides along
    with `metadata` rather than being listed per view — a tab that forgot it would render
    its breakups with the wrong columns and nothing would fail.
    """
    page = (REPO_ROOT / "app" / "dashboard" / "[view]" / "page.tsx").read_text(encoding="utf-8")
    assert '.in("key", [...spec.scalars, "metadata", "detail_columns"])' in page
    # A breakup is read from the build the figure came from. The policies hold a viewer
    # to the current build but let an admin read every build, so an unfiltered query
    # would merge them — the defect that once rendered 1,188 rows for a 396-row section.
    panel = (REPO_ROOT / "app" / "dashboard" / "[view]" / "detail.tsx").read_text(encoding="utf-8")
    assert '.eq("build_id", buildId)' in panel
    # The prefix leads the primary key and is what `can_read_prefix` checks a grant on.
    assert '.eq("prefix", prefix)' in panel


def test_every_drill_down_on_the_web_app_reaches_a_breakup():
    """A drill-down template is a string, and no compiler checks where it points.

    `views.ts` names the key each clickable figure opens — `{stock_detail_key}` where the
    pipeline precomputed one, `LLSCHEDULE|{bucket}` where it is composed. Rename a field
    upstream and the column keeps compiling and starts opening nothing, silently. So
    every template is resolved against every row of the published payload and checked
    against the drill-downs the same build produced.
    """
    import shutil

    # The checker is a Node script because it evaluates the real view declarations rather
    # than parsing them. Where there is no Node there is no web app to check either.
    node = shutil.which("node")
    if not node:
        return

    result = subprocess.run(
        [str(node), str(REPO_ROOT / "tools" / "check_detail_keys.mjs"),
         str(REPO_ROOT / "data.json")],
        capture_output=True, text=True, cwd=REPO_ROOT,
    )
    assert result.returncode == 0, result.stdout + result.stderr
    # The headline cards are the only breakups nothing opens, and deliberately: the fact
    # strip that replaced them states figures rather than offering them.
    assert "Not opened from any table: BALANCE" in result.stdout


def test_the_web_apps_customer_tracker_carries_the_same_columns_as_the_page():
    """The two dashboards answer the same question and must not answer it differently.

    The web app's tracker had drifted: it had dropped the three piece columns, was showing
    the CTL pool in tonnes where the page shows it in pieces, and had added a WIP column
    the page does not carry — so a customer's line read differently depending on which URL
    you opened. These are the page's own thirteen headings, in its order.
    """
    template = (SKILL_ROOT / "assets" / "dashboard_template.html").read_text(encoding="utf-8")
    head = re.search(r'<table id="customerTable">\s*<thead><tr>(.*?)</tr>', template, re.S)
    assert head, "the static page's customer table has a header row"
    expected = re.findall(r"<th[^>]*>([^<]+)</th>", head.group(1))
    assert expected == [
        "OEM", "Customer", "SKU / CTL bucket", "Plant",
        "Schedule Qty", "Dispatch Qty", "Balance Qty",
        "Schedule MT", "Dispatch MT", "Balance MT",
        "CTL stock NOS", "LL stock MT", "Avg month sales",
    ], expected

    views = (REPO_ROOT / "app" / "dashboard" / "[view]" / "views.ts").read_text(encoding="utf-8")
    block = re.search(r"const customerLineColumns = \(\): Column\[\] => \[(.*?)\n\];", views, re.S)
    assert block, "the shared line columns are declared in one place"
    # A column is written either as a shorthand call — `mt("field", "Label")` — or as the
    # object the shorthands build. Both forms are read, in source order, so the test
    # compares the order of the headings and not just the set of them.
    labels = [
        found.group(1) or found.group(2)
        for found in re.finditer(r'\(\s*"[^"]*",\s*"([^"]+)"|label: "([^"]+)"', block.group(1))
    ]
    assert labels == expected, labels
    # The columns the page does not carry stay off it.
    for absent in ("WIP MT", "Open bal MT", "Over disp MT", "Bucket"):
        assert f'"{absent}"' not in block.group(1), absent


def test_the_customer_tab_asks_for_a_customer_before_it_answers():
    """396 lines across sixteen customers answers nobody's question.

    The static page renders nothing until a customer is chosen, and its three tables all
    read the same choice. The web app holds that choice in the URL so the server does the
    narrowing and the tables cannot disagree with each other — the same rule the tonnes /
    pieces switch follows.
    """
    views = (REPO_ROOT / "app" / "dashboard" / "[view]" / "views.ts").read_text(encoding="utf-8")
    customer = views[views.index("customerView: {"):views.index("meghView: {")]
    assert 'param: "customer"' in customer
    # All three of the customer's tables narrow on the one selection.
    assert customer.count("pickFields:") == 3
    assert 'pickFields: { customer: "customer_display" }' in customer  # lines, CRFH book
    assert 'pickFields: { customer: "customer" }' in customer          # the sales history

    page = (REPO_ROOT / "app" / "dashboard" / "[view]" / "page.tsx").read_text(encoding="utf-8")
    # Narrowed before derived, so a join inside `flatten` sees one customer's rows.
    assert page.index("table.pickFields") < page.index("table.flatten")
    # A table narrowing on a selector nobody has answered — as against the summary you
    # choose from, which names none — stays off the page rather than showing everything.
    assert "const waiting = awaiting.some((p) => table.pickFields?.[p.param]);" in page
    assert "if (waiting) return false;" in page


def test_uploading_is_a_tab_that_cannot_be_granted_to_a_viewer():
    """It sits in the strip with the eleven views and is deliberately not one of them.

    A `dashboard_views` row can be granted, and a grant is the wrong shape for this: who
    may load a dump is a role. So the tab is rendered on the role and the page redirects
    on it, while the database refuses the write regardless — the control not being drawn
    has never been access control on this project.
    """
    views_sql = (REPO_ROOT / "supabase" / "migrations"
                 / "20260809173652_identity_and_view_grants.sql").read_text(encoding="utf-8")
    assert "uploadView" not in views_sql and "'upload'" not in views_sql

    layout = (REPO_ROOT / "app" / "dashboard" / "layout.tsx").read_text(encoding="utf-8")
    assert 'href="/dashboard/upload"' in layout
    assert "{canUpload && (" in layout

    page = (REPO_ROOT / "app" / "dashboard" / "upload" / "page.tsx").read_text(encoding="utf-8")
    assert 'user.role !== "admin" && user.role !== "uploader"' in page

    # The old address still answers, so a bookmark or a mail does not dead-end.
    legacy = (REPO_ROOT / "app" / "upload" / "page.tsx").read_text(encoding="utf-8")
    assert 'redirect("/dashboard/upload")' in legacy

    # A sheet missing a column the pipeline reads is not written at all.
    uploader = (REPO_ROOT / "app" / "dashboard" / "upload"
                / "uploader.tsx").read_text(encoding="utf-8")
    assert "unreadable" in uploader
    assert "disabled={!ready || busy || unreadable.length > 0}" in uploader

    # The write path the browser uses has to exist in a migration, not only in the live
    # database: a Supabase branch or a fresh project has to be able to accept an upload.
    migrations = "\n".join(
        path.read_text(encoding="utf-8")
        for path in (REPO_ROOT / "supabase" / "migrations").glob("*.sql")
    )
    for definition in ("function public.promote_upload(", "function public.is_uploader()",
                       '"uploaders create batches"',
                       '"uploaders add rows to their own pending batch"'):
        assert definition in migrations, definition


def test_the_browser_adapters_match_the_pipelines_own_registry():
    """`lib/dumps/adapters.ts` is generated, and nothing was checking that it still was.

    Both the generator and the generated header claim a test fails when the checked-in
    copy goes stale, and until now none did — so a slot added to `sources.py` would have
    gone on being invisible to the uploader with nothing to say so. Now the claim is true.
    """
    import shutil

    python = shutil.which("python3") or sys.executable
    result = subprocess.run(
        [python, str(REPO_ROOT / "tools" / "generate_adapters.py"), "--check"],
        capture_output=True, text=True, cwd=REPO_ROOT,
    )
    if result.returncode != 0 and "ModuleNotFoundError" in result.stderr:
        # No pandas on this interpreter: the generator imports the pipeline. Nothing to
        # prove here rather than a false failure.
        return
    assert result.returncode == 0, result.stdout + result.stderr


def test_the_uploader_names_a_sheets_columns_the_way_the_read_will():
    """The upload page's column check is only worth having if the names agree.

    A dump whose headers have moved used to be accepted, look fine, and fail hours later
    inside the refresh as a `KeyError` naming one column — by which time whoever has the
    file has gone home. The page now says which declared columns it found before anything
    is stored, and that means naming them the way the read names them: blanks numbered,
    duplicates suffixed, the column window applied, trailing empty unnamed columns
    dropped. Written twice, in `columns.ts` and in `sources.py`, so it is compared.

    A cell holding a single space is the case that matters. It is a value to pandas and
    looks like nothing to a reader, and one of them in the last column of the TVSM sheet
    is the difference between twenty columns and twenty-one.

    Unlike the other Node checks this one parses .xlsx, so it needs SheetJS installed.
    The refresh runner has Python and Node and deliberately no `npm ci` — it builds no
    web app — so there the check is skipped rather than failed. It ran and passed on the
    machine the change was made on; run it there after touching either side.
    """
    import shutil
    import tempfile

    node = shutil.which("node")
    if not node or not (REPO_ROOT / "node_modules" / "xlsx").exists():
        return

    with tempfile.TemporaryDirectory() as tmp:
        seen_path = Path(tmp) / "columns.json"
        result = subprocess.run(
            [str(node), str(REPO_ROOT / "tools" / "check_upload_columns.mjs"),
             str(REPO_ROOT / "dumps"), str(seen_path)],
            capture_output=True, text=True, cwd=REPO_ROOT,
        )
        assert result.returncode == 0, result.stdout + result.stderr
        seen = json.loads(seen_path.read_text(encoding="utf-8"))

    assert len(seen) >= 20, f"only {len(seen)} slots parsed"

    pipeline = load_refresh_module()
    # Loading the pipeline puts its own directory on the import path, which is what makes
    # the sibling registry importable here by name.
    import sources

    extra = sources.slots_for_families(
        pipeline.ORDER_BOOK_SHEETS, pipeline.PRICING_SHEETS, pipeline.SIGNOFF_SHEETS
    )
    slots = dict(sources.SLOTS)
    slots.update(extra)
    excel = sources.ExcelSources(REPO_ROOT / "dumps", extra_slots=extra)

    for slot, got in sorted(seen.items()):
        assert not got["missing"], f"{slot}: {got['missing']}"
        if got["positional"]:
            continue
        spec = slots[slot]
        sheet = got["sheet"] if spec.sheet is sources.CALLER_NAMES_THE_SHEET else None
        frame = excel.frame(slot, sheet=sheet)
        assert [str(c) for c in frame.columns] == got["columns"], slot
        # A preview that shows nothing is not a preview.
        assert got["sampleRows"] > 0, slot


def test_the_quarterly_calculation_is_a_drill_down_and_leaves_megh_out():
    """The CN/DN working is thousands of billing lines, so it cannot be a section.

    A tab fetches a section whole and the page caps one at 3,000 rows; a quarter is around
    eight thousand billing lines across all customers. So it is written as `detail_rows`
    under a `RECO` prefix, fetched by key for the picked customer and quarter alone, and
    the page prefetches it because a copy has to be built inside the click that asked for
    it — the clipboard is not writable after an await.

    Megh Steel is left out on purpose rather than by omission: it is billed per kilogram
    against a conversion rate and its reconciliation does not follow the contract formula.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    block = script[script.index("reco_lines = trend["):]
    block = block[:block.index("reco_quarters = sorted(")]
    assert 'trend["segment"].eq(TREND_SEGMENT_DIRECT)' in block, "Megh must be left out"
    assert 'stock_details[f"RECO|{customer}|{quarter}"]' in block
    # The three quantity columns the working is built on, and the trap in the third.
    for field in ('"qty_nos"', '"qty_m"', '"qty_kg"', '"billing_rate"', '"sales_unit"'):
        assert field in block, field
    assert 'float(r["sales_mt"] or 0) * 1000' in block, "qty in KG comes off Quantity"
    # An invoice number that arrives as a float matches nothing in the customer's file.
    assert 'whole_number_text(r["Billing  Document Number"])' in block

    # April starts the financial year, so January to March close the one already named.
    refresh = load_refresh_module()
    assert refresh.financial_quarter("2026-01") == "Q4 FY26"
    assert refresh.financial_quarter("2025-10") == "Q3 FY26"
    assert refresh.financial_quarter("2026-04") == "Q1 FY27"
    assert refresh.financial_quarter("2026-07") == "Q2 FY27"
    assert refresh.financial_quarter(None) is None

    views = (REPO_ROOT / "app" / "dashboard" / "[view]" / "views.ts").read_text(encoding="utf-8")
    pricing = views[views.index("pricingView: {"):views.index("stockView: {")]
    # Only the quarters the build holds billing for, and the customer comes from the
    # tab's selector rather than a control of its own.
    assert "reco_quarters" in pricing and 'param: "customer"' in pricing
    assert 'kind: "calculation" as const' in pricing
    assert "prefetchDetails:" in pricing

    page = (REPO_ROOT / "app" / "dashboard" / "[view]" / "page.tsx").read_text(encoding="utf-8")
    # Paged: PostgREST caps a response at a thousand rows and one customer's quarter can
    # be well past that, and a claim short by a third looks complete.
    assert "async function fetchDetail(" in page
    assert "spec.prefetchDetails?.(picks, scalars)" in page

    copies = (REPO_ROOT / "app" / "dashboard" / "[view]" / "copies.ts").read_text(encoding="utf-8")
    for column in ("QTY IN NOS", "QTY IN M", "QTY IN KG", "BILLING RATE", "PO RATE",
                   "DIFFERENCE OF PRICE", "REJECTION", "FINAL QTY", "CN/DN VALUE", "CN/DN"):
        assert f'"{column}"' in copies, column
    # Rejection is blank, never zero: the dashboard cannot know it, and a zero reads as
    # "none" rather than "not filled in yet".
    assert "const finalQty = billedQty;" in copies


class _FakeRest:
    """PostgREST, as far as `PostgresSources` can tell.

    Enough to prove the absorption path builds the right requests without a database:
    the conflict target and resolution that make the insert additive, the grand-total row
    never reaching the table, and the batch being stamped so it is absorbed exactly once.
    """

    def __init__(self, batches, grids):
        self.batches = batches
        self.grids = grids
        self.inserted = []
        self.insert_calls = []
        self.updated = []

    def select(self, table, query=""):
        if table == "raw_batches":
            if "absorbed_at=is.null" in query:
                return [b for b in self.batches if b.get("absorbed_at") is None]
            return self.batches
        if table == "raw_rows":
            batch_id = query.split("batch_id=eq.")[1].split("&")[0]
            offset = int(query.split("offset=")[1].split("&")[0])
            limit = int(query.split("limit=")[1].split("&")[0])
            cells = self.grids[batch_id]
            return [{"seq": i, "row": row}
                    for i, row in enumerate(cells)][offset:offset + limit]
        if table == "tsl_sales":
            return []
        raise AssertionError(f"unexpected read of {table}")

    def insert(self, table, rows, *, chunk=2000, returning=False,
               prefer=None, on_conflict=None):
        self.insert_calls.append({"table": table, "prefer": prefer,
                                  "on_conflict": on_conflict, "rows": len(rows)})
        self.inserted.extend(rows)

    def update(self, table, query, patch, *, returning=False):
        self.updated.append((table, query, patch))


def test_the_refresh_absorbs_before_it_reads_the_ledger():
    """A method nothing calls is a method that does nothing.

    This is the defect that shipped: `absorb_sales` was written, unit-tested against a
    fake client, and never wired in. Every test passed — the one below included — because
    each exercised a piece, and the missing thing was the line joining them. The offline
    `dumps/` run could not catch it either: `ExcelSources` assembles the ledger from files
    and has no call site to forget. So the call itself is what is asserted here, and its
    position: absorption after the pipeline would be as useless as none at all.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_from_supabase.py").read_text(
        encoding="utf-8"
    )
    assert "absorb_sales()" in script, (
        "nothing fills tsl_sales, so the ledger reads empty and the pipeline dies on "
        "the first sales column it touches"
    )
    assert script.index("absorb_sales()") < script.index("pipeline.main("), (
        "the ledger is the pipeline's input; absorbing after it runs is too late"
    )


def test_the_refresh_reads_the_dump_tables_not_the_cell_grid():
    """The whole point of putting the dumps in tables, asserted at the one line that does it.

    `TableSources` and `PostgresSources` serve the same slots and differ in where they
    read them from, so swapping one for the other changes nothing that raises and
    everything that matters: read off the grid, the pipeline sees the newest upload of the
    transfer dump — the month in progress — instead of every line the ledger has absorbed
    since 8 July. Measured at the switch: 220 transfer lines against 255, 1,898 MT against
    2,129, and 416 MT in transit against 439.

    Asserted on the source text for the same reason
    `test_the_refresh_absorbs_before_it_reads_the_ledger` is: nothing else in the suite
    reaches this line, because the offline `dumps/` run uses `ExcelSources` and never
    constructs either of them.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_from_supabase.py").read_text(
        encoding="utf-8"
    )
    assert "sources_module.TableSources(" in script, (
        "the refresh is reading the stored cell grid again, so every accumulating dump "
        "is back to whatever the newest upload happens to carry"
    )
    assert "sources_module.PostgresSources(client" not in script


def test_every_slot_with_a_table_can_be_read_back_out_of_it():
    """A table nothing can read back is a copy of the dump, not a source of it.

    The pipeline addresses every column by the header the file wrote — `CUSTOMER  CD`
    with two spaces, `Material\xa0No` with a non-breaking one, `Chamferring ` with a
    trailing space — and the views name their columns in snake case because an Excel
    header is not an identifier. `config/dump_columns.json` is the mapping between the
    two, and a slot missing from it is a slot the refresh silently falls back to reading
    off the cell grid.

    The two exemptions are exemptions on purpose. `contract:*` keeps no table at all — it
    is read with `header=None` because its quarters are column offsets, so there are no
    names to map. The five sales slots pour into one ledger keyed on the invoice line,
    where `sales_q1` and today's dump are the same rows by design; `sales_ledger()` reads
    that table and no individual slot can be read back out of it.
    """
    load_refresh_module()
    import sources

    manifest = sources.dump_columns()
    for slot in all_input_slots():
        spec = sources.table_for(slot)
        if spec is None or slot in sources.SALES_LEDGER_SLOTS:
            continue
        if slot == "schedule_supplement":
            # Retired: the owner's v15 workbook carries every customer, so there is no
            # supplement to read and nothing in `dumps/` to name its columns from.
            continue
        assert slot in manifest, (
            f"{slot!r} keeps its rows in {spec.table} and nothing says how to read them "
            f"back, so the refresh would quietly read the cell grid instead"
        )
        assert manifest[slot]["table"] == spec.table
        assert manifest[slot]["columns"], f"{slot!r} maps no columns"


def test_a_view_never_types_an_identifier_as_a_number():
    """The one thing a view gets wrong that no read can put right.

    A column is typed from one file in `dumps/`, once. `yf65.xlsx` there writes the
    accounting document number as a *number*, so the view was written `dump_numeric` —
    and the uploaded `yf65.XLSX` writes the same column as `0071029066`. The padding is
    stripped in SQL, before anything can read it, and it reached the page: the overdue
    drill-down showed `DP 36388067` where the invoice is `DP 0036388067`, in a document
    people paste into a mail. The order book and both sign-off sheets had the same fault
    on ten more columns — `000000000110102155` read as 110102155, `00001` as 1, `056`
    as 56.

    So the identifier columns are typed by name rather than by whichever file was to hand,
    and this holds the generated migration to it. `tools/compare_table_sources.py --drift`
    is what finds the next one, against the uploads actually held.
    """
    sys.path.insert(0, str(REPO_ROOT / "tools"))
    import generate_dump_views

    migration = snapshot_view_migration()
    for name in sorted(generate_dump_views.TEXT_COLUMNS):
        numeric = f"public.dump_numeric(r.row -> "
        for line in migration.splitlines():
            if line.strip().endswith(f"as {name},") or line.strip().endswith(f"as {name}"):
                assert numeric not in line, (
                    f"{name!r} is an identifier and this view reads it as a number, "
                    f"which strips the zero padding SAP writes and cannot be undone"
                )


def test_a_number_revives_out_of_a_view_and_a_padded_code_does_not():
    """`dump_text` collapses the number 788 and the text `0788`; this is what separates them.

    Both come back as text, and the same stock column holds both — 2,158 rows of one and
    298 of the other. Left as text, every numeric column that shares a sheet with a padded
    code changes type under the pipeline; converted blindly, `0788` becomes 788 and a code
    that meant a plant stops joining to anything.
    """
    load_refresh_module()
    import sources

    assert sources.revive_number("788") == 788
    assert sources.revive_number("2.042") == 2.042
    assert sources.revive_number("-3") == -3
    # A code, and it stays one. Every shape SAP actually pads in.
    assert sources.revive_number("0788") == "0788"
    assert sources.revive_number("00001") == "00001"
    assert sources.revive_number("000000000003502027") == "000000000003502027"
    # Not the shortest text of any float, so not a number a spreadsheet wrote.
    assert sources.revive_number("2.0420") == "2.0420"
    assert sources.revive_number("1e3") == "1e3"
    assert sources.revive_number("+3") == "+3"
    assert sources.revive_number("ERW 1-FC") == "ERW 1-FC"
    assert sources.revive_number(None) is None
    assert sources.revive_number(6) == 6


def test_a_date_revives_out_of_a_table_and_the_text_of_one_does_not():
    """JSON has no date, so `_storable` writes a Timestamp as its own ISO text.

    Coming back it is indistinguishable from a cell that held that text — and a whole
    column of dates read as text is not a small difference: every `.dt` accessor
    downstream stops working and a date comparison silently becomes a string comparison,
    which orders `2026-1-9` after `2026-10-9`. The round trip is what tells them apart.
    """
    load_refresh_module()
    import pandas as pd
    from datetime import time as clock

    import sources

    assert sources.revive_moment("2026-07-14T00:00:00") == pd.Timestamp("2026-07-14")
    assert sources.revive_moment("17:55:00") == clock(17, 55)
    # A cell that really did hold this text: no time on it, so it is not what
    # `.isoformat()` writes and it is left exactly as it came.
    assert sources.revive_moment("2026-07-14") == "2026-07-14"
    assert sources.revive_moment("TUB-O-N-25.4") == "TUB-O-N-25.4"
    assert sources.revive_moment(None) is None


def test_zmat_is_keyed_on_the_row_and_not_just_the_code_and_plant():
    """zmat has no natural key, so the row's own content is the third part of it.

    Keyed on `(material_code, plant)` alone the table dropped 1,104 real rows. They looked
    like noise — an end finish and a surface finish swapped, a specification written `10`
    on one row and `010` on the next — and they are not: the pipeline identifies a
    material by its code, its description and an attribute key built from both diameters,
    the thickness, the specification and both finishes, so two rows that pair calls the
    same are two different materials to everything downstream.

    Measured, reading the pipeline off the table against off the sheet: nine stock rows
    stopped resolving to a bucket, the STR plan lost two lines, and a long-length SKU's
    signed-off tonnage read 4.925 MT against a true 7.205.
    """
    load_refresh_module()
    import pandas as pd

    import sources

    def row(**overrides):
        line = {column: None for _, column, _ in sources.ZMAT_COLUMNS}
        line.update({"Column1": "3501105", "PLANT": "0788",
                     "MATERIAL DESCRIPTION": "TUB-O-N-25.4"})
        line.update(overrides)
        return line

    frame = pd.DataFrame([
        row(**{"MATERIAL END FINISH": "PE", "MATERIAL SURFACE FINISH": "AW"}),
        row(**{"MATERIAL END FINISH": "AW", "MATERIAL SURFACE FINISH": "PE"}),
        # Byte-identical to the first, and this one really is a repeat.
        row(**{"MATERIAL END FINISH": "PE", "MATERIAL SURFACE FINISH": "AW"}),
    ])
    keyed = sources.zmat_keys(frame)
    assert len(keyed) == 2, (
        "two rows differing in end and surface finish are two materials to the pipeline; "
        "collapsing them is what cost nine stock resolutions"
    )
    assert set(sources.TABLES["zmat"].key) == {"material_code", "plant", "row_digest"}
    # The sheet's own order, kept, because the pipeline deduplicates again with
    # `keep="first"` and "first" has to mean the row the sheet wrote first.
    assert sources.TABLES["zmat"].read_order == ("source_seq",)
    assert list(keyed["source_seq"]) == [0, 1]


def test_an_empty_sales_ledger_stops_the_refresh():
    """And says which read came back empty, rather than which column went missing.

    Unguarded, this surfaced as `KeyError: 'CUSTOMER  CD'` three hundred lines downstream
    — a column that is in every sales file and was never the problem. The frame had no
    columns at all.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    guard = script[script.index("    sales_ledger = src.sales_ledger()"):]
    guard = guard[:guard.index("published_month")]
    assert "if sales_ledger.empty:" in guard, "an empty ledger must not reach derive_sales"
    assert "absorb_sales" in guard, (
        "the message has to name what fills the ledger, or it explains nothing"
    )


def test_absorbing_a_sales_batch_adds_only_lines_the_ledger_has_never_seen():
    """The ledger's whole contract, exercised without a database.

    Three things have to be true of the request, and none of them is visible from the
    frame it was built from. The insert has to name a conflict target *and* a resolution
    — PostgREST treats a resolution without a target as an ordinary insert and raises on
    the first duplicate key, which would turn a re-sent day into a failed refresh. The
    key has to arrive as text. And the batch has to be stamped absorbed, because that
    stamp is the only thing standing between a superseded batch and being pruned with
    its lines never recorded.
    """
    load_refresh_module()
    import sources

    columns = list(sources.SALES_COLUMNS)
    def cell_row(values):
        return [values.get(c) for c in columns]

    cells = [
        columns,
        cell_row({"Billing  Document Number": 4731002954.0, "Billing Item": 10.0,
                  "BILLING  DATE": "2026-08-01", "Quantity": 1000.0,
                  "MATERAIL NUMBER": "3907863", "CUSTOMER  CD": "943209",
                  "DESP P LANT": "0789"}),
        # The sheet's own grand total: no document, no date, and the column's sum.
        cell_row({"Quantity": 968438.483}),
    ]
    batch = {"id": "b1", "slot": "sales", "sheet": None, "row_count": len(cells),
             "original_filename": "sales.xlsx", "absorbed_at": None}
    client = _FakeRest([batch], {"b1": cells})

    source = sources.PostgresSources(client)
    absorbed = source.absorb_sales(log=lambda _message: None)

    assert absorbed == {"b1": 1}, "the grand-total row is not a line"
    assert len(client.inserted) == 1
    line = client.inserted[0]
    assert line["billing_document"] == "4731002954", "float out of pandas, text into the key"
    assert line["billing_item"] == "10"
    assert line["billing_month"] == "2026-08"
    assert line["customer_code"] == "943209"
    # The material code must survive as the text it is; inferred as a float, pandas drops
    # its last digit and 3907863 becomes 3907860.
    assert line["row"]["MATERAIL NUMBER"] == "3907863"
    assert line["source_batch"] == "b1"

    call = client.insert_calls[0]
    assert call["on_conflict"] == "billing_document,billing_item", (
        "without a conflict target the resolution is ignored and a re-sent day raises"
    )
    assert "resolution=ignore-duplicates" in call["prefer"]

    assert client.updated == [
        ("raw_batches", "id=eq.b1", {"absorbed_at": "now()"})
    ], "an unstamped batch is absorbed again, or pruned before it ever is"


def _transfer_cells(rows):
    """A transfer grid: the header, then one cell row per dict given."""
    columns = ["MATERAIL NUMBER", "Material   Description", "DESP P LANT", "CUSTOMER  CD",
               "Invoice Type", "GR DATE", "Quantity", "qty in no",
               "Billing  Document Number", "Billing Item", "BILLING  DATE"]
    return [columns] + [[row.get(c) for c in columns] for row in rows]


def test_a_transfer_batch_that_is_really_the_sales_dump_is_refused():
    """The mis-sent file stops being self-correcting the moment the table accumulates.

    The daily mail has more than once carried a copy of the sales dump under the transfer
    filename — the 27 July set arrived byte-identical, 3,990 rows and 234 columns both.
    Under the snapshot model the next upload cleared it. In a table that accumulates,
    those sales invoices would sit in the transfer ledger permanently, and nothing could
    take them out again. So the batch is refused before a row of it is written.

    The test is exact rather than a proportion because the fault is exact: a real extract
    is `Tax Inv Transfer` and `D.Challan Transfer` throughout, and a sales extract carries
    `Tax Inv Sale IGST`/`SGST` and matches nothing.

    And it is skipped rather than raised. One file in the wrong slot must not stop the
    good ones and must certainly not stop the sales ledger the pipeline reads next, so a
    refusal takes the bad batch out of the run and leaves everything else in it.
    """
    load_refresh_module()
    import sources

    bad = _transfer_cells([
        {"Invoice Type": "Tax Inv Sale IGST", "Billing  Document Number": 4731002954.0,
         "Billing Item": 10.0, "DESP P LANT": 789.0},
        {"Invoice Type": "Tax Inv Sale SGST", "Billing  Document Number": 4731002955.0,
         "Billing Item": 10.0, "DESP P LANT": 789.0},
    ])
    good = _transfer_cells([
        {"Invoice Type": "Tax Inv Transfer", "Billing  Document Number": 4470110377.0,
         "Billing Item": 10.0, "DESP P LANT": 788.0, "CUSTOMER  CD": 8406.0},
    ])
    batches = [
        {"id": "bad", "slot": "transfers", "sheet": None, "row_count": len(bad),
         "original_filename": "transfer.xlsx", "absorbed_at": None},
        {"id": "good", "slot": "transfers", "sheet": None, "row_count": len(good),
         "original_filename": "transfer_correct.xlsx", "absorbed_at": None},
    ]
    client = _FakeRest(batches, {"bad": bad, "good": good})

    said = []
    absorbed = sources.PostgresSources(client).absorb(log=said.append)

    assert absorbed == {"good": 1}, (
        "the good batch behind a refused one must still be absorbed"
    )
    assert len(client.inserted) == 1, "no sales line may reach the transfer ledger"
    assert client.inserted[0]["billing_document"] == "4470110377"
    assert client.updated == [("raw_batches", "id=eq.good", {"absorbed_at": "now()"})], (
        "a refused batch must stay un-absorbed, or the refusal is recorded as success "
        "and the dump is pruned unstored"
    )
    assert any("REFUSED" in message and "transfer.xlsx" in message for message in said), (
        "a refusal nobody is told about is a dump that quietly never arrives"
    )


def test_absorbing_a_transfer_batch_keys_it_on_the_invoice_line():
    """The same key as sales, the same total-row rule, and plants that will join.

    Three things this has to get right and none is visible from the frame. The key halves
    arrive as floats — the sheet's grand total is what makes pandas read them that way —
    and must be stored as the text they are. The grand-total row itself carries no
    billing document and is dropped for having none rather than deduplicated. And both
    plant columns go through `plant_code`, because this dump writes `788.0` where stock
    writes `0788` and zmat writes `788`, and a stock-transfer plan joining the raw values
    would report a code as not extended when it plainly is.
    """
    load_refresh_module()
    import sources

    cells = _transfer_cells([
        {"Invoice Type": "Tax Inv Transfer", "Billing  Document Number": 4470110377.0,
         "Billing Item": 10.0, "DESP P LANT": 788.0, "CUSTOMER  CD": 8406.0,
         "MATERAIL NUMBER": "000000000003501105", "BILLING  DATE": "2026-07-11",
         "GR DATE": "2026-07-20", "Quantity": 1000.0},
        # Despatched but not yet received: no GR date is the in-transit flag.
        {"Invoice Type": "D.Challan Transfer", "Billing  Document Number": 4470110378.0,
         "Billing Item": 10.0, "DESP P LANT": 56.0, "CUSTOMER  CD": 4731.0,
         "MATERAIL NUMBER": "000000000003501106", "BILLING  DATE": "2026-07-15",
         "Quantity": 2000.0},
        # The sheet's own grand total: no invoice type, no document, the column's sum.
        {"Quantity": 3000.0},
    ])
    batch = {"id": "b1", "slot": "transfers", "sheet": None, "row_count": len(cells),
             "original_filename": "transfer.xlsx", "absorbed_at": None}
    client = _FakeRest([batch], {"b1": cells})

    absorbed = sources.PostgresSources(client).absorb(log=lambda _message: None)

    assert absorbed == {"b1": 2}, "the grand-total row is not a despatch"
    first, second = client.inserted
    assert first["billing_document"] == "4470110377", "float out of pandas, text into the key"
    assert first["billing_item"] == "10"
    assert first["billing_month"] == "2026-07"
    assert first["gr_date"] == "2026-07-20"
    assert first["source_plant"] == "788" and first["receiving_plant"] == "8406"
    assert first["row"]["MATERAIL NUMBER"] == "000000000003501105", (
        "a material code that arrived zero-padded text must not come back a number"
    )
    assert second["gr_date"] is None, "an absent goods receipt is what in-transit means"
    assert second["source_plant"] == "56"

    call = client.insert_calls[0]
    assert call["table"] == "tsl_transfers"
    assert call["on_conflict"] == "billing_document,billing_item", (
        "without a conflict target the resolution is ignored and a re-sent day raises"
    )
    # Which resolution travels with that target is a separate claim, and belongs with the
    # reason for it: see `test_a_transfer_line_is_allowed_to_stop_being_in_transit`.
    assert client.updated == [("raw_batches", "id=eq.b1", {"absorbed_at": "now()"})]


def test_a_transfer_line_is_allowed_to_stop_being_in_transit():
    """The one place the transfer ledger must not copy the sales ledger.

    A billed sale is finished when it is billed, so `tsl_sales` keeps the first version of
    a line and ignores every later one. A transfer is not: it stays *in transit* until the
    receiving plant posts a goods receipt, and `GR DATE` therefore fills in on a dump sent
    days after the one that first carried the line. Measured on the four transfer dumps
    held, 227 of 1,088 lines did exactly that.

    Keep-first froze all 227 as in transit permanently — the table reported 445 against a
    true 218 — and the error was invisible, because every individual line looked right.
    So the resolution is `merge-duplicates`, and the natural key still does the job it was
    chosen for: a re-sent day cannot double count, it can only correct.
    """
    load_refresh_module()
    import sources

    assert sources.TABLES["transfers"].on_conflict == "replace", (
        "keep-first freezes GR DATE, and a line in transit stays in transit for ever"
    )
    assert sources.TABLES["sales"].on_conflict == "keep", (
        "a billed sale is finished when it is billed; letting a later dump rewrite it "
        "would let a re-cut extract silently move a closed month"
    )

    cells = _transfer_cells([
        {"Invoice Type": "Tax Inv Transfer", "Billing  Document Number": 4470110597.0,
         "Billing Item": 1.0, "DESP P LANT": 788.0, "CUSTOMER  CD": 8406.0,
         "BILLING  DATE": "2026-07-11", "GR DATE": "2026-07-20"},
    ])
    batch = {"id": "b1", "slot": "transfers", "sheet": None, "row_count": len(cells),
             "original_filename": "transfer.xlsx", "absorbed_at": None}
    client = _FakeRest([batch], {"b1": cells})
    sources.PostgresSources(client).absorb(log=lambda _message: None)

    assert "resolution=merge-duplicates" in client.insert_calls[0]["prefer"]
    assert client.insert_calls[0]["on_conflict"] == "billing_document,billing_item", (
        "PostgREST ignores a resolution given without a conflict target, so the two must "
        "travel together or the second dump raises on the first duplicate key"
    )


def test_bucketing_is_keyed_on_the_material_code_and_not_on_the_bucket():
    """The slot's declared key column is a label, and keying on it loses the master.

    `ReadSpec.key_column` is documented as the column that says what a row is *about* —
    it is what the uploader shows first in a preview, not a natural key. For `Bucketting`
    it is `Bucket`, the size family a code belongs to, and it is shared: 167 distinct
    buckets over 1,538 rows, with `12.7-0-1.6-ERW 1-PE` alone covering 27 material codes.
    Keyed on it the table would hold 167 rows and lose nine tenths of the mapping, and
    nothing about the insert would fail.
    """
    load_refresh_module()
    import sources

    assert sources.TABLES["bucketting"].key == ("material_code",), (
        "keyed on Bucket, 1,538 rows of mapping collapse to 167"
    )
    assert sources.SLOTS["bucketting"].key_column == "Bucket", (
        "if the read spec stops calling it Bucket this test has lost its point"
    )


def test_a_re_bucketed_code_updates_and_a_code_dropped_from_the_workbook_survives():
    """Both halves of what the owner asked of the mapping masters.

    Accumulating is the easy half: a code the newest workbook does not mention is not
    thereby deleted. The other half is that the newest workbook *wins* where they
    disagree, because codes do get re-bucketed and an OEM's spelling does get corrected.
    Keep-first would mean neither ever landed, and `bucket_assignments` — which exists to
    override the master — would become the only way to correct it.
    """
    load_refresh_module()
    import sources

    for slot in ("bucketting", "oem_key"):
        assert sources.TABLES[slot].on_conflict == "replace", slot

    # `Bucketting` is read through a column window, `U:AF` — columns 21 to 32 of the
    # sheet, so 20 empty ones come first. Building the grid without them is how a test of
    # this slot passes while reading nothing at all.
    lead = [None] * 20
    columns = ["Material Codes", "OD", "ID", "Thickness", "Length", "Grade", "FC/PE",
               "Annealed", "Bucket", "LL or CTL", "CTL Bucket", "kG/nos"]
    cells = [
        lead + ["ignored"] * len(columns),   # header=1, so row 0 is above the names
        lead + columns,
        lead + [2426342.0, 11, None, 1.5, 0.263, "WG", "PE", "N",
                "11--1.5-CEW-NON FIN", "CTL", "11--1.5-CEW-NON FIN-0.263", 1.2],
        # The sheet's trailing padding: empty in the file, and not a bucket mapping.
        lead + [None] * len(columns),
    ]
    batch = {"id": "b1", "slot": "bucketting", "sheet": "Bucketting",
             "row_count": len(cells), "original_filename": "rm_tracker_model.xlsx",
             "absorbed_at": None}
    client = _FakeRest([batch], {"b1": cells})
    absorbed = sources.PostgresSources(client).absorb(log=lambda _m: None)

    assert absorbed == {"b1": 1}, "the padding rows carry no code and are not mappings"
    row = client.inserted[0]
    assert row["material_code"] == "2426342", (
        "read as float64 because the column has blanks, and stored raw it would be "
        "'2426342.0' and join to nothing on the stock or sales side"
    )
    assert row["bucket"] == "11--1.5-CEW-NON FIN"
    call = client.insert_calls[0]
    assert call["table"] == "dump_bucketing"
    assert call["on_conflict"] == "material_code"
    assert "resolution=merge-duplicates" in call["prefer"], (
        "ignore-duplicates would freeze a code in the first bucket it was ever seen in"
    )


def test_zmat_is_keyed_on_the_code_and_the_plant_it_is_extended_at():
    """The plant is half the key, and it is the half the table exists for.

    zmat is a material × plant extract — the same code once per plant it is extended at,
    otherwise identical. Keyed on the code alone it is 57,478 rows instead of 64,074 and
    answers nothing about where a material can be made, which is precisely what a
    stock-transfer plan asks of both the sending and the receiving plant.
    """
    load_refresh_module()
    import sources

    assert sources.TABLES["zmat"].key[:2] == ("material_code", "plant")
    assert sources.SLOTS["zmat"].key_column == "Column1", (
        "the read spec's key column is the code alone, which is why the table cannot "
        "take its key from there"
    )


def test_zmat_is_deduplicated_in_frame_order_and_not_by_the_database():
    """No column combination in this file is unique, so somebody has to choose.

    480 rows are byte-identical repeats of another row. Left to
    `resolution=ignore-duplicates`, PostgREST would resolve against whatever happened to
    share a 2,000-row chunk, so which of two identical rows survived would move the day
    the file gains a row above them. Deduplicating here makes it frame order, which is
    the order the file was written in.

    **What is deduplicated changed on 14 August, and this test changed with it.** It used
    to assert that two rows sharing a code and a plant collapsed to the first — including
    the pair below, which differ in their description. That was wrong, and wrong in a way
    only visible from the far end: `(Column1, PLANT)` leaves 1,104 rows over beyond the
    byte-identical ones, and the pipeline identifies a material by its code, its
    description and an attribute key built from both diameters, the thickness, the
    specification and both finishes. Two rows that pair calls the same are two different
    materials to all of it. Collapsed, nine stock rows stopped resolving to a bucket, the
    STR plan lost two lines, and a long-length SKU's signed-off tonnage read 4.925 MT
    against a true 7.205.

    So the third part of the key is a digest of the row's own content. Byte-identical
    repeats still collapse — that is all a digest can do — and anything the file wrote
    differently is kept.
    """
    load_refresh_module()
    import sources
    import pandas as pd

    frame = pd.DataFrame([
        # Same code at two plants: both are real and both must survive.
        {"Column1": 3406677, "PLANT": 56,  "MATERIAL DESCRIPTION": "TUB-A"},
        {"Column1": 3406677, "PLANT": 788, "MATERIAL DESCRIPTION": "TUB-A"},
        # Same code and plant, different description: two materials, both kept.
        {"Column1": 249132, "PLANT": 789, "MATERIAL DESCRIPTION": "TUB-B"},
        {"Column1": 249132, "PLANT": 789, "MATERIAL DESCRIPTION": "TUB-B-CORRECTED"},
        # And a byte-identical repeat of the row above it, which is not a material.
        {"Column1": 249132, "PLANT": 789, "MATERIAL DESCRIPTION": "TUB-B-CORRECTED"},
    ])
    keyed = sources.zmat_keys(frame)
    assert len(keyed) == 4, "only the byte-identical repeat is a repeat"
    assert list(keyed["plant"]) == ["56", "788", "789", "789"]
    assert list(keyed["MATERIAL DESCRIPTION"]) == [
        "TUB-A", "TUB-A", "TUB-B", "TUB-B-CORRECTED",
    ], "frame order, first wins"

    # Twice over the same frame gives the same answer, which is what "deterministic" means
    # here and what `ignore-duplicates` could not promise.
    assert list(sources.zmat_keys(frame)["MATERIAL DESCRIPTION"]) == \
        list(keyed["MATERIAL DESCRIPTION"])


def test_one_bad_cell_does_not_stop_the_material_master_loading():
    """A typed column is what makes this table fit, and it is also what makes it brittle.

    `OUTER DIAMETER OF MATERIAL` is the text `o` on exactly one row of the 65,178 held — a
    typo for zero — and it failed the insert for the whole 2,000-row chunk it travelled
    in: `invalid input syntax for type numeric: "o"`. Coerced to null rather than widening
    the column, because a diameter that cannot be filtered on numerically is worth less
    than one missing value, and nothing is lost either way: the batch is still in
    `raw_rows` exactly as uploaded.
    """
    load_refresh_module()
    import sources
    import pandas as pd

    frame = sources.zmat_keys(pd.DataFrame([
        {"Column1": 3406677, "PLANT": 56, "OUTER DIAMETER OF MATERIAL": "o"},
        {"Column1": 3406678, "PLANT": 56, "OUTER DIAMETER OF MATERIAL": 41.28},
    ]))
    records = sources.zmat_records(frame, "b1")
    assert records[0]["outer_diameter"] is None, "a cell that is not a number is not one"
    assert records[1]["outer_diameter"] == 41.28


def test_material_codes_from_every_dump_canonicalise_to_the_same_value():
    """`000000000003501105` and `3501105` are one material, and both are written.

    SAP holds a material number zero-padded to eighteen characters and shows it unpadded,
    and which reaches a dump is decided per extract. Measured on what is stored: the
    transfer dump pads all 1,088 of its lines and WIP all 693, stock and the bucketing
    master pad none, and the sales ledger pads 6,539 of 22,419 — the daily dump and the
    quarterly archives disagree with each other. Joined raw, 0 of 1,088 transfer lines
    reached a bucket; joined on this, 837 do.
    """
    load_refresh_module()
    import sources

    for written in ("000000000003501105", "3501105", 3501105, 3501105.0, " 3501105 "):
        assert sources.material_code(written) == "3501105", written
    assert sources.material_code(None) is None
    assert sources.material_code("") is None
    # A code that is not digits is left exactly as it came, and one that is all zeros
    # keeps a zero rather than becoming nothing.
    assert sources.material_code("6 M CODE") == "6 M CODE"
    assert sources.material_code("0000") == "0"


def test_plant_codes_from_every_dump_canonicalise_to_the_same_value():
    """`0788`, `788` and `788.0` are one plant, and only one of them is written anywhere.

    Measured on the live uploads: `stock.Plant` holds `0788` and `789` in the same column,
    `zmat.PLANT` holds `788` and `789`, the sales extract holds `056` and `0788`, and the
    transfer dump holds floats. The 8406 stock-transfer plan asks whether a material code
    is extended at both plants, so a join across these is the whole feature — and on the
    raw values it matches a subset and calls the rest absent.
    """
    load_refresh_module()
    import sources

    for written in ("0788", "788", 788, 788.0, " 788 "):
        assert sources.plant_code(written) == "788", written
    assert sources.plant_code("056") == "56"
    assert sources.plant_code(None) is None
    assert sources.plant_code("") is None
    # A plant code that is not digits is left exactly as it came: one this does not
    # recognise is not one it should rewrite.
    assert sources.plant_code("HOSUR") == "HOSUR"
    # And zero does not vanish when its padding is stripped.
    assert sources.plant_code("0000") == "0"


def test_pruning_refuses_a_batch_whose_rows_are_not_in_a_table_yet():
    """Pruning must not outrun absorption, for every slot that accumulates.

    A dump uploaded and left unrefreshed for a fortnight would otherwise be deleted with
    its rows never recorded, and nothing downstream would report a gap — the months would
    simply be short.

    Derived from the registry rather than spelled out, because the failure this guards
    against is somebody adding an accumulating slot and not coming back here. A slot named
    in `TABLES` as accumulating and absent from the clause is exactly that.
    """
    load_refresh_module()
    import sources

    # The clause as the *latest* migration to redefine the function leaves it: an earlier
    # one still carries the version it replaced, and asserting against that would pass
    # while production ran something else.
    written = [path.read_text(encoding="utf-8")
               for path in sorted((REPO_ROOT / "supabase" / "migrations").glob("*.sql"))]
    latest = [text for text in written
              if "create or replace function public.prune_uploads" in text][-1]
    clause = latest[latest.index("create or replace function public.prune_uploads"):]
    clause = clause[:clause.index("returning 1")]

    assert "absorbed_at is not null" in clause
    for slot, spec in sorted(sources.TABLES.items()):
        if spec.mode != "accumulating":
            continue
        # `sales%` covers the five sales slots between them; anything else is named.
        covered = f"'{slot}'" in clause or (slot.startswith("sales") and "'sales%'" in clause)
        assert covered, (
            f"{slot!r} accumulates but pruning does not wait for it — a dump uploaded and "
            f"left unrefreshed for a fortnight is deleted with its rows never stored"
        )


def test_sales_to_megh_opens_months_that_add_to_the_figure_above_them():
    """The published month's column must total the very figure that was clicked.

    Every other breakup on the Megh tab splits one month's figure by where the material
    is. This one answers a different question — how the size has sold over time — and a
    history beside a monthly number is only honest if the two meet somewhere. They meet
    in the published month's column, and that is what makes the panel checkable.

    Two things have to hold for it, and both have been got wrong on this dashboard
    before. Each month column must say `add: True`, because `kind: "mt"` is not summable
    on its own — it is also weight per metre on the price build-up — and a layout relying
    on it renders a footer of blanks. And `MEGHSALES` must stay out of `AVERAGE_BY_MONTH`,
    which closes a panel on an average month: right where the *rows* are months, wrong
    here where they are material codes and the months are columns.
    """
    refresh = load_refresh_module()

    layout = refresh.megh_sales_detail_columns(["2026-07", "2026-08"])
    assert [c["label"] for c in layout] == [
        "Material code", "Description", "Jul 26", "Aug 26", "Total",
    ]
    months = [c for c in layout if c["field"].startswith("m_")]
    assert months and all(c.get("add") is True for c in months), (
        "a month column that is not summable leaves the footer blank"
    )
    assert layout[-1]["kind"] == "qty", "the total column is the row's own quantity"

    detail = (REPO_ROOT / "app" / "dashboard" / "[view]" / "detail.tsx").read_text(
        encoding="utf-8"
    )
    average = detail[detail.index("const AVERAGE_BY_MONTH"):]
    average = average[:average.index(")")]
    assert "MEGHSALES" not in average, (
        "MEGHSALES closes on a total, not an average month: its rows are codes, not months"
    )

    # The figure is guarded on the months behind it, not on the month's own tonnage —
    # a size that sold in March and nothing since is exactly the row worth opening.
    views = (REPO_ROOT / "app" / "dashboard" / "[view]" / "views.ts").read_text(
        encoding="utf-8"
    )
    # From the figure to the end of the `drill(...)` that wraps it. Closing on the
    # dedented `),` rather than the first one, which belongs to the `mt(...)` inside it.
    block = views[views.index('mt("sales_mt", "Sales to Megh MT")'):]
    block = block[:block.index("\n          ),")]
    # The guard is `drill`'s last argument, so it is the block's last line. Checking the
    # position rather than mere presence is what tells it apart from the field name in
    # the `mt(...)` the drill wraps.
    guard = block.strip().splitlines()[-1].strip()
    assert guard == '"sales_months",', (
        f"the history must be guarded on the months behind it, not on {guard}"
    )


def test_meghs_conversion_cost_model_reproduces_the_owners_own_workbook():
    """Megh's claim is a different calculation, and this is the half that is certain.

    An ancillary buys a tube and is billed for it, so its claim is the contract price
    against the billed price. Megh *converts*: what it should have been billed is what it
    can sell on at, less what converting costs it. The owner's workbook solves that by
    hand — a goal seek driving cost minus realisation to zero — and because the cost is
    linear in the rate being solved for there is a closed form.

    Pinned against the owner's own Q4 FY26 and Q1 FY27 workbooks when it was written: all
    1,093 and 1,211 TVS lines respectively. The figures below are three of those lines,
    kept here so the model cannot drift without the test saying so.
    """
    refresh = load_refresh_module()

    # Four real lines, one per case the workbook contains — long length and cut piece,
    # a month of carrying and a week of it. Every figure is read off the owner's own
    # `Column1`, not computed here: a test that recomputes the thing it is testing is a
    # test of nothing.
    assert round(refresh.megh_ex_jsr_rate(78650, 30, True), 2) == 71981.97
    assert round(refresh.megh_ex_jsr_rate(75900, 7, True), 2) == 69879.22
    assert round(refresh.megh_ex_jsr_rate(75350, 7, False), 2) == 71034.33
    assert round(refresh.megh_ex_jsr_rate(76850, 0, False), 2) == 72636.58
    # Nothing to solve for is None, not zero: zero is a rate and this is an absence.
    assert refresh.megh_ex_jsr_rate(0, 30, True) is None
    assert refresh.megh_ex_jsr_rate(None, 30, True) is None

    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    # The claim columns stay empty until Megh's base-price master is a held input.
    block = script[script.index("megh_reco_lines = trend_all["):]
    block = block[:block.index("megh_reco_quarters = sorted(")]
    for blank in ('"proposed_rate_per_kg": None', '"difference_per_kg": None',
                  '"cn_dn_value": None', '"cn_dn": None'):
        assert blank in block, blank
    # And the charged price is taken off the gross invoice value, as the owner's own
    # workbook does — the two disagree on a handful of lines and the gross is what wins.
    assert 'gross / MEGH_RECO_GST_FACTOR / kg' in block

    # The fourth conversion code, spelt the way the OEM key spells it.
    assert refresh.CONVERSION_AGENT_OEM_BY_CODE["943213"] == "Rane"


def test_a_quarters_working_says_which_plants_it_stands_on():
    """Two of the archived sales extracts hold the southern plants only.

    `sales_q4.xlsx` and `sales_q1.xlsx` carry no Jamshedpur and no Khopoli, where
    `sales_jul.xlsx` and the daily dump carry every plant. For Megh Steel alone that is
    125 of the 200 invoices in Q1 FY27. A claim built off a quarter like that is missing
    a third of itself and looks exactly like a complete one, so every working prints the
    plants it covers.

    What it must not do is guess *why* a plant is absent. A plant missing from one
    quarter and present in another is either a gap in the extract or a plant that shipped
    nothing, and 4318 really did stop — so the coverage is stated and not judged.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    block = script[script.index("quarter_plants = {}"):]
    block = block[:block.index("# ---- Megh Steel's own quarterly working")]
    assert '"missing_plants": missing' in block
    assert '"missing_share": round(share, 4)' in block
    assert '"short"' not in block and '"complete"' not in block, (
        "coverage is stated, not judged: a boolean here would be a guess"
    )

    copies = (REPO_ROOT / "app" / "dashboard" / "[view]" / "copies.ts").read_text(encoding="utf-8")
    assert "function coverageLine(" in copies
    # Both workings carry it, not just the new one.
    assert copies.count("coverageLine(ctx,") == 2


def test_the_browser_prices_a_sku_the_way_the_pipeline_does():
    """The contract formula is written twice and both copies have to agree.

    The price column, the build-up behind it and the PO rate the quarterly calculation
    quotes are all recomputed in the browser, so a reader who corrects a SKU's operations
    sees the price move without waiting for a refresh. That is a second implementation of
    the pipeline's arithmetic, and two implementations drift silently: the page would show
    one number, tomorrow's build another, and each would look right on its own.

    `tools/check_pricing_formula.mjs` runs the TypeScript against the pipeline's own
    output and compares every price and every build-up line. It also checks that the key
    an override is recorded against names exactly one SKU — it did not until the governed
    bucket was made part of it.

    Run here against `data.json`, which is the frozen 7 August oracle and therefore
    predates both the published base per metre and the bucket in the key. So this run
    proves the prices and the key; **the build-up comparison needs a build newer than the
    oracle** and the tool says how many it skipped. Run it against a fresh build after any
    change to the formula on either side:

        python .claude/skills/refresh-tvsm-dashboard/scripts/refresh_dashboard.py \\
            --input-dir dumps --output-dir /tmp/build --as-of <date>
        node tools/check_pricing_formula.mjs /tmp/build/data.json
    """
    import shutil

    node = shutil.which("node")
    if not node:
        return

    result = subprocess.run(
        [str(node), str(REPO_ROOT / "tools" / "check_pricing_formula.mjs"),
         str(REPO_ROOT / "data.json")],
        capture_output=True, text=True, cwd=REPO_ROOT,
    )
    assert result.returncode == 0, result.stdout + result.stderr
    assert "The page and the build agree" in result.stdout, result.stdout
    # A pass that checked nothing is not a pass: name the counts it must reach.
    assert re.search(r"agree: 1,?125 prices", result.stdout), result.stdout
    assert "375 keyable SKUs" in result.stdout, result.stdout


def test_an_owner_correction_overrides_the_schedules_operation_flags():
    """The pricing tab's operation flags are right most of the time and wrong some of it.

    Every SKU where this view disagrees with a customer's own reconciliation is an
    operation question — a fin-cut ladder charged on a `PE` line the schedule does not
    flag, a fin cut charged per `FC` that the customer waived. So the correction is an
    **override** and not a fallback: a fallback would fix none of those cases, because in
    every one of them the flags said something.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    # Read at the start of a run and written back, so a clean clone rebuilds the same way.
    assert "PRICING_OVERRIDES_FILE" in script
    assert "def load_pricing_overrides(" in script
    assert (SKILL_ROOT / "config" / "pricing_overrides.json").exists()

    block = script[script.index("def pricing_operation_override("):]
    block = block[:block.index("def norm_code(")]
    assert 'overrides or {}).get("operations", {}).get(key)' in block

    applied = script[script.index('operations["Fin cut"] = PRICING_OPERATION_RATES'):]
    applied = applied[:applied.index("operations_per_m = (")]
    # The override replaces the set rather than adding to it.
    assert "operations = {" in applied and "for name in corrected" in applied

    # And the key names one SKU: customer, governed bucket, material code, length.
    key = script[script.index("def pricing_override_key("):]
    key = key[:key.index("def pricing_operation_override(")] if \
        "def pricing_operation_override(" in key[key.index("\n"):] else key[:2000]
    assert "row['bucket']" in script
    assert 'f"PRICEBUILD|{row[\'customer\']}|{row[\'bucket\']}|{row[\'material_code\']}"' in script


def test_the_trend_groups_a_customers_sap_names_and_closes_on_an_average_month():
    """The sales file names a ship-to, not a customer, and a total is not a rate.

    Twenty-six SAP spellings cover about thirteen customers — Rajsriya arrives under six
    — so a selector offering the raw names asks the reader to know which spelling is the
    plant they meant. Each row therefore also carries the Helper Customer its SAP code
    belongs to, and a code claimed by two Helper Customers keeps its raw name rather than
    being guessed at. The table then closes on the tonnage over the months that actually
    moved, because a window total sitting beside seven month columns reads as a monthly
    figure.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    block = script[script.index("trend_customer_skus = []"):]
    block = block[:block.index("trend_customer_skus.sort")]
    for field in ('"customer_group"', '"customer"', '"months_active"',
                  '"avg_active_month_mt"', '"avg_active_month_nos"'):
        assert field in block, field
    # Averaged over the months that moved, never over the window.
    assert "round(total_mt / len(months), 3) if months else None" in block
    # A shared code is not resolved: only codes with exactly one Helper Customer are.
    assert "if len(displays) == 1" in script
    assert 'trend["customer_key"].map(unique_display_by_code).fillna(' in script

    views = (REPO_ROOT / "app" / "dashboard" / "[view]" / "views.ts").read_text(encoding="utf-8")
    trend = views[views.index("trendView: {"):views.index("pricingView: {")]
    # Two selectors, the ship-to narrowing inside the customer.
    assert 'param: "customer"' in trend and 'param: "shipto"' in trend
    assert 'within: "customer"' in trend
    assert 'pickFields: { customer: "customer_group", shipto: "customer" }' in trend
    # The window total is off the table; the average and its denominator are on it.
    skus = trend[trend.index('key: "trend_customer_skus"'):trend.index('key: "trend_customer_sku_history"')]
    assert "unitTotal(ctx.unit)" not in skus
    assert 'cntNoTotal("months_active", "Months")' in skus
    assert "averageOver:" in skus

    # The average still divides the window total, which is on the row but not a column,
    # so the totals row must sum that field itself rather than read it out of `totals`.
    table = (REPO_ROOT / "app" / "dashboard" / "[view]" / "table.tsx").read_text(encoding="utf-8")
    assert "num(get(rows[i], averageOver.totalField))" in table

    # A unit switch that rebuilt the URL from scratch would clear the customer.
    page = (REPO_ROOT / "app" / "dashboard" / "[view]" / "page.tsx").read_text(encoding="utf-8")
    assert 'withParam(query, "unit", option)' in page


def test_the_crfh_book_is_split_out_for_the_two_customers_that_ask_for_it():
    """Marathwada and Sri Balaji Gear read their CRFH range as its own book.

    Mixed into the tube lines it distorts every total on the table above it, which is why
    the static page lifts it out. The split is on the bucket, so it needs no customer list
    to maintain: only those two have CRFH rows, and the table hides itself for everyone
    else rather than showing an empty sheet with an explanation.
    """
    views = (REPO_ROOT / "app" / "dashboard" / "[view]" / "views.ts").read_text(encoding="utf-8")
    assert 'String(row.bucket ?? "").toUpperCase().includes("CRFH")' in views
    assert "flatten: (rows) => rows.filter((row) => !isCrfh(row))" in views, "tube lines exclude it"
    assert "flatten: (rows) => rows.filter(isCrfh)" in views, "the book is only CRFH"
    assert "hideWhenEmpty: true" in views


def test_the_customer_tab_may_read_the_history_it_shows():
    """An ungranted section returns no rows, and a page cannot tell that from no sales.

    The history is the same section the trend tab reads whole; the customer tab reads it
    one customer at a time. Without this row a reader granted only the customer tab would
    see an empty history and no reason for it.
    """
    migrations = REPO_ROOT / "supabase" / "migrations"
    granted = "".join(
        path.read_text(encoding="utf-8") for path in sorted(migrations.glob("*.sql"))
    )
    assert "('trend_customer_sku_history', 'customerView')" in granted


def test_the_two_sent_documents_still_read_exactly_as_the_static_pages():
    """These are pasted into WhatsApp and into the dispatch team's sheet.

    A difference of a decimal place or a thousand separator is a wrong figure in someone
    else's document and nothing about it looks wrong, so the equivalence is re-proved
    rather than remembered. It matters most right after a change to what the table hands
    the buttons: splitting the CRFH book into its own table would otherwise have dropped
    those lines out of Marathwada's and Sri Balaji Gear's dispatch plans silently.
    """
    import shutil

    node = shutil.which("node")
    if not node:
        return

    result = subprocess.run(
        [str(node), str(REPO_ROOT / "tools" / "compare_copy_formats.mjs")],
        capture_output=True, text=True, cwd=REPO_ROOT,
    )
    assert result.returncode == 0, result.stdout + result.stderr
    assert "32 of 32 documents are byte-identical" in result.stdout, result.stdout


def test_an_owner_assignment_overrides_the_master_and_governs_the_code():
    """The point of the Missing mappings tab is that answering it changes something.

    On the static page the assignment dropdowns wrote to an object in memory: the choice
    vanished on reload and moved no tonnage on any tab. The decision is now kept against
    the material code, and the pipeline applies it at the one place a code becomes a
    bucket — `material_bucket`, which every tracker, stock frame and queue joins through.

    It is an override rather than a fallback on purpose. A code the master resolves to the
    *wrong* bucket is exactly the case somebody is correcting, and a fallback would
    silently ignore them.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    resolve = script.index('material_bucket = zmat.groupby("material_key")')
    override = script.index('material_bucket.loc[code] = bucket')
    assert resolve < override, "the owner's answer is applied after the master's"
    # Applied once, at the join every view reads. A second place to override is a second
    # place to forget.
    assert script.count("material_bucket.loc[code] = bucket") == 1

    refresh = load_refresh_module()
    # A caller-supplied set is used as given: nothing is read and nothing is written, so
    # a test and a clean-clone rebuild are both reproducible.
    import inspect

    assert "assignments" in inspect.signature(refresh.main).parameters


def test_the_assignments_the_build_used_are_committed_beside_it():
    """A clean clone with no credentials has to reproduce the published build.

    The decisions live in `bucket_assignments`, which the browser writes and no clean
    clone can reach. So the run that reads them writes them into the repository, and a
    rebuild without credentials reads that file instead. Without this the reproducibility
    check would start failing the first time anybody assigned anything.
    """
    refresh = load_refresh_module()
    committed = SKILL_ROOT / "config" / "bucket_assignments.json"
    assert committed.exists(), "the file the pipeline falls back to is committed"
    held = json.loads(committed.read_text(encoding="utf-8"))
    assert "assignments" in held and isinstance(held["assignments"], dict)

    # Reading with the refresh turned off must not touch the network or the file.
    before = committed.read_bytes()
    assert refresh.load_bucket_assignments(refresh=False) == held["assignments"]
    assert committed.read_bytes() == before

    # And the table it reads is not scoped to a build — which is the whole point.
    migrations = "".join(
        path.read_text(encoding="utf-8")
        for path in sorted((REPO_ROOT / "supabase" / "migrations").glob("*.sql"))
    )
    assignments = migrations[migrations.index("create table public.bucket_assignments"):]
    assignments = assignments[:assignments.index(");")]
    assert "build_id" not in assignments, "an assignment must outlive the build it was made on"
    assert "primary key (scope, material_code)" in assignments


def test_every_scope_the_browser_can_write_is_one_the_pipeline_reads():
    """A queue that saves and changes nothing is worse than a queue with no box on it.

    `megh_sku` was exactly that for a fortnight: the cell wrote it, the cell read it back,
    and the Megh tab never moved, because the Megh plan keys off its own mapping and not
    off `material_bucket`. The queue looked answered while the tonnage stayed missing.

    So the rule is a closed loop: for each of the three spaces, the browser may write it,
    the database accepts it, and the pipeline applies it. Adding a fourth scope without
    somewhere to apply it fails here rather than on the tab three weeks later.
    """
    script = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")

    # `bucket` — the material master's own resolution, overridden at the one join.
    assert 'owner_assignments = (assignments or {}).get("bucket", {})' in script

    # `megh_sku` — applied after the two routes to a plan key have been tried, so the
    # owner's answer wins over both rather than only filling a hole they left.
    fallback = script.index('frame["material_key"].map(code_to_vsm_key)')
    sku_override = script.index('frame["material_key"].map(sku_assignments)')
    family = script.index('frame["vsm_family"] = frame["vsm_key"].map(key_family)')
    assert 'sku_assignments = (assignments or {}).get("megh_sku", {})' in script
    assert fallback < sku_override < family, (
        "the assignment overrides the lookup, and the family is read off the result"
    )

    # `ctl_bucket` — RFD alone, because RFD alone resolves through the master's own CTL
    # column and so cannot be reached by a `bucket` assignment.
    resolved = script.index('rfd["ctl_bucket"] = rfd["material_key"].map(direct["CTL Bucket"])')
    ctl_override = script.index("listed.map(ctl_assignments)")
    assert 'ctl_assignments = (assignments or {}).get("ctl_bucket", {})' in script
    assert resolved < ctl_override, "the owner's answer is applied after the master's"


def test_the_assignable_spaces_agree_everywhere():
    """One list of scopes, held in five places, and a typo in any of them loses a decision.

    The cell posts a scope, the route admits it, the database's check constraint stores it
    and the pipeline reads it back by the same string. A value saved under a scope nothing
    reads is not an error anywhere — it saves, it shows, and it does nothing.

    The fifth place is the pipeline, and it is the one this test used to take on trust.
    `oem` was added to the other four on 17 August and withdrawn the same night for
    exactly that: the constraint stored it, the route admitted it, the cell reported it
    saved, and `oem_map` never looked. `megh_sku` sat that way for a fortnight. So the
    last assertion here reads the pipeline itself, and a scope that nothing consumes
    fails rather than shipping as a box that lies about having saved something.
    """
    scopes = {"bucket", "megh_sku", "ctl_bucket", "oem"}

    migrations = "".join(
        path.read_text(encoding="utf-8")
        for path in sorted((REPO_ROOT / "supabase" / "migrations").glob("*.sql"))
    )
    # The latest statement about the constraint is the one in force.
    constraint = migrations.rindex("check (scope in (")
    stated = migrations[constraint:migrations.index("))", constraint)]
    assert {s.strip(" '") for s in stated.split("(")[-1].split(",")} == scopes

    route = (REPO_ROOT / "app" / "api" / "assign" / "route.ts").read_text(encoding="utf-8")
    admitted = route[route.index("const SCOPES"):route.index("\n", route.index("const SCOPES"))]
    assert {s for s in scopes if f'"{s}"' in admitted} == scopes

    views = (REPO_ROOT / "app" / "dashboard" / "[view]" / "views.ts").read_text(encoding="utf-8")
    offered = set(re.findall(r'scope: "([a-z_]+)"', views))
    assert offered <= scopes, f"a queue offers a space nothing accepts: {offered - scopes}"

    # And the fifth place: the pipeline reads every one of them by that same string.
    pipeline = (SKILL_ROOT / "scripts" / "refresh_dashboard.py").read_text(encoding="utf-8")
    read = {s for s in scopes if f'.get("{s}", {{}})' in pipeline}
    assert read == scopes, (
        f"assignable but never read by the pipeline: {sorted(scopes - read)}. "
        "A decision filed under one of these saves, reads back, and moves nothing."
    )


def test_the_assign_cell_takes_an_answer_the_master_does_not_carry_yet():
    """A dropdown of governed buckets had this queue backwards.

    A row is on the Missing mappings tab *because* nothing governs it, so the bucket that
    belongs on it is routinely one `Bucketting` has never held — and a `<select>` of what
    already exists cannot express that. It is a text field now: the governed list is still
    suggested, a value outside it is saved, and the cell flags that the master needs the
    same addition rather than refusing the answer.
    """
    cell = (REPO_ROOT / "app" / "dashboard" / "[view]" / "assign.tsx").read_text(encoding="utf-8")
    # Past the header comment, which says what this used to be and why it is not that.
    code = cell.split("*/", 1)[1]
    assert "<select" not in code, "the choice is not limited to what already exists"
    assert "<datalist" in cell, "the governed list is still offered, as a suggestion"
    # Saved, then flagged — never blocked. If this ever guards the fetch instead, an owner
    # would be unable to record the one answer the queue exists to collect.
    assert "not in the master yet" in cell
    warn = cell.index("not in the master yet")
    assert cell.index("await fetch(\"/api/assign\"") < warn

    # A save per keystroke would post thirty times for one bucket, and the last few would
    # race. The two things that end an edit are what save it.
    assert "onBlur=" in cell and 'event.key === "Enter"' in cell

    # And the two guarantees the control has always carried, unchanged by any of it.
    assert "setValue(previous)" in cell
    # The wording moved when Apply mappings arrived — the property did not: a saved cell
    # says the figures still need a rebuild, never implies they already moved.
    assert "press Apply mappings to rebuild" in cell


def test_only_the_admin_may_assign_and_the_database_is_what_says_so():
    """A control that is merely not drawn is not access control.

    The field is rendered for an admin alone, but anyone can post to the route. So the
    write goes through the caller's own client and the policy decides, and the refusal is
    reported rather than swallowed — a cell that silently does nothing is worse than one
    that says it was refused.
    """
    migrations = "".join(
        path.read_text(encoding="utf-8")
        for path in sorted((REPO_ROOT / "supabase" / "migrations").glob("*.sql"))
    )
    for policy in ("admin assigns", "admin reassigns", "admin unassigns"):
        assert f'create policy "{policy}" on public.bucket_assignments' in migrations
    assert 'create policy "read assignments" on public.bucket_assignments' in migrations

    route = (REPO_ROOT / "app" / "api" / "assign" / "route.ts").read_text(encoding="utf-8")
    assert "supabaseServer()" in route, "the caller's client, so the policy decides"
    assert "supabaseAdmin" not in route, "the service role would bypass the policy it relies on"
    assert "42501" in route, "a policy refusal is named as one"

    cell = (REPO_ROOT / "app" / "dashboard" / "[view]" / "assign.tsx").read_text(encoding="utf-8")
    # The screen never shows a decision the database did not take.
    assert "setValue(previous)" in cell
    # And never implies the figures have moved, because they have not until a rebuild —
    # which is what the Apply mappings button the cell points at actually runs.
    assert "press Apply mappings to rebuild" in cell


def test_an_applied_mapping_rebuilds_under_the_builds_own_date():
    """Apply mappings republishes the same dumps under the date they arrived on.

    The one failure the refresh exists to stop is an unchanged dump re-stamped with a
    fresh date — ageing, days-overdue and the schedule month all read the as-of, so a
    today-dated rebuild of yesterday's dumps ages every figure at once and looks exactly
    like real movement. The daily refresh may default to today because its dumps really
    did just arrive; a mapping applied the next morning must not. So the button sends
    the current build's own as-of, the route validates and forwards it, and the workflow
    reads it off the dispatch. Three files, one date, and losing any link re-stamps.
    """
    apply = (REPO_ROOT / "app" / "dashboard" / "[view]" / "apply.tsx").read_text(encoding="utf-8")
    assert "as_of: asOf" in apply, "the button must send the build's as-of, not nothing"

    route = (REPO_ROOT / "app" / "api" / "refresh" / "route.ts").read_text(encoding="utf-8")
    assert r"^\d{4}-\d{2}-\d{2}$" in route, "the route validates the date before dispatching"
    assert "as_of: asOf" in route, "the route forwards it in the client_payload"

    workflow = (REPO_ROOT / ".github" / "workflows" / "refresh.yml").read_text(encoding="utf-8")
    assert "github.event.client_payload.as_of" in workflow, (
        "the workflow must read the dispatched as-of; without this the payload is "
        "carried and ignored, and the rebuild silently stamps today"
    )

    # And the page hands the button the build's own date, not the browser's.
    page = (REPO_ROOT / "app" / "dashboard" / "[view]" / "page.tsx").read_text(encoding="utf-8")
    assert "ApplyMappings asOf={String(scalars.metadata.as_of)}" in page


def test_a_remark_outlives_the_build_and_applies_at_once():
    """Why an invoice is late is written from the panel and kept outside the build.

    Two things separate it from a bucket assignment, and both are easy to lose. It cannot
    be scoped to a build — the next refresh replaces every drill-down row and their key
    ends in a sort position. And it applies immediately, because it feeds no figure, so
    the cell must not carry the assign cell's "applies at the next refresh" caveat.
    """
    migrations = "".join(
        path.read_text(encoding="utf-8")
        for path in sorted((REPO_ROOT / "supabase" / "migrations").glob("*.sql"))
    )
    remarks = migrations[migrations.index("create table public.invoice_remarks"):]
    remarks = remarks[:remarks.index(");")]
    assert "build_id" not in remarks, "a remark must outlive the build it was made on"
    assert "invoice_no  text primary key" in remarks, "held against the invoice, not the row"

    # Whoever may read the overdue tab may remark on it, and the database is what says so.
    for policy in ("read overdue remarks", "record a remark", "revise a remark",
                   "withdraw a remark"):
        assert f'create policy "{policy}" on public.invoice_remarks' in migrations
    assert migrations.count("public.can_read_view('overdueView')") == 5

    route = (REPO_ROOT / "app" / "api" / "remark" / "route.ts").read_text(encoding="utf-8")
    assert "supabaseServer()" in route, "the caller's client, so the policy decides"
    assert "supabaseAdmin" not in route, "the service role would bypass the policy it relies on"
    assert "42501" in route, "a policy refusal is named as one"

    cell = (REPO_ROOT / "app" / "dashboard" / "[view]" / "remark.tsx").read_text(encoding="utf-8")
    # The screen never shows a remark the database did not take.
    assert "setValue(previous)" in cell
    # A remark moves no figure, so there is nothing to wait for and nothing to caveat.
    assert "applies at the next refresh" not in cell.split("*/", 1)[1]

    # And the panel holds them apart from the rows it caches by build, or a save would be
    # undone on screen the next time the same figure was opened.
    panel = (REPO_ROOT / "app" / "dashboard" / "[view]" / "detail.tsx").read_text(encoding="utf-8")
    assert "invoice_remarks" in panel
    assert "cache.set(slot, rows)" in panel and "cache.set(slot, remarks" not in panel
